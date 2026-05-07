import io
import os
import csv
import math
import shutil
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
from PySide6.QtUiTools import QUiLoader
from PySide6.QtCore import Qt, QDateTime, QTime, Signal
from PySide6.QtWidgets import (QVBoxLayout, QPushButton, QDateTimeEdit, QDialog, QFileDialog,
    QLabel, QFrame, QWidget, QHBoxLayout, QCalendarWidget, QListWidget, QLineEdit)
from utils.common.alertas import mostrar_mensaje
from utils.common.rutasarchivos import resource_path
from utils.common.metodosGenerales import MetodosGenerales
from modules.empresa.softwareconfiguracion import SoftwareConfiguracion
from modules.empresa.empresaconfiguracion import EmpresaConfiguracion
from controllers.DatosController import DatosController

class TimeWheel(QListWidget):
    def __init__(self, limit, parent=None):
        super().__init__(parent)
        self.setFixedWidth(35)
        self.setFixedHeight(120)
        self.setUniformItemSizes(True)
        self.setVerticalScrollMode(QListWidget.ScrollPerPixel)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        for i in range(limit):
            self.addItem(f"{i:02d}")
            self.item(i).setTextAlignment(Qt.AlignCenter)
        self.setStyleSheet("QListWidget { border: 1px solid #ddd; background: white; color: #333; font-size: 11px; } QListWidget::item:selected { background: #0078d7; color: white; }")

class DateTimePickerPopup(QDialog):
    def __init__(self, parent=None, initial_dt=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.Popup | Qt.FramelessWindowHint)
        self.main_frame = QFrame(self)
        self.main_frame.setStyleSheet("QFrame { background: white; border: 1px solid #ccc; border-radius: 6px; }")
        layout = QVBoxLayout(self.main_frame)
        layout.setContentsMargins(5, 5, 5, 5)
        
        body = QHBoxLayout()
        self.calendar = QCalendarWidget()
        self.calendar.setFixedSize(280, 210) 
        
        # --- ESTILO ULTRA-FINO DEL CALENDARIO ---
        self.calendar.setStyleSheet("""
            /* Barra de navegación */
            QCalendarWidget QWidget#qt_calendar_navigationbar {
                background-color: #ffffff;
                border-bottom: 1px solid #f2f2f2;
            }

            /* Botones generales de la barra superior */
            QCalendarWidget QToolButton {
                color: #333333;
                background-color: transparent;
                border: none;
                height: 25px;
            }

            /* --- SELECTOR DE MES (COMBO) --- */
            QCalendarWidget QToolButton#qt_calendar_monthbutton {
                font-size: 11px;
                font-weight: bold;
                padding-right: 12px; /* Espacio para nuestra flechita */
                padding-left: 5px;
                margin-right: 2px;
            }

            /* Personalización de la flechita del combo de meses */
            QCalendarWidget QToolButton#qt_calendar_monthbutton::menu-indicator {
                subcontrol-origin: padding;
                subcontrol-position: center right;
                right: 2px; /* Separación del borde derecho */
                top: 0px;   /* Ajuste vertical para alineación perfecta */
                width: 8px;  /* Flechita mucho más pequeña */
                height: 8px;
            }

            /* --- SELECTOR DE AÑO --- */
            QCalendarWidget QToolButton#qt_calendar_yearbutton {
                font-size: 11px;
                font-weight: bold;
                margin-left: 2px;
                padding: 0 5px;
            }

            /* Flechas laterales (Mes anterior/siguiente) */
            QCalendarWidget QToolButton#qt_calendar_prevmonth, 
            QCalendarWidget QToolButton#qt_calendar_nextmonth {
                width: 24px;
                border-radius: 12px;
                qproperty-iconSize: 14px;
            }
            
            QCalendarWidget QToolButton:hover {
                background-color: #f5f5f5;
                border-radius: 4px;
            }

            /* Menú desplegable de meses */
            QCalendarWidget QMenu {
                background-color: white;
                color: #333;
                selection-background-color: #0078d7;
                border: 1px solid #eeeeee;
            }

            /* Grilla de días y números */
            QCalendarWidget QWidget { alternate-background-color: #ffffff; }
            QCalendarWidget QAbstractItemView:enabled {
                color: #444;
                selection-background-color: #0078d7;
                font-size: 11px;
            }
            QCalendarWidget QAbstractItemView:disabled { color: #d0d0d0; }
        """)

        # Configuración de visibilidad
        self.calendar.setHorizontalHeaderFormat(QCalendarWidget.SingleLetterDayNames)
        self.calendar.setVerticalHeaderFormat(QCalendarWidget.NoVerticalHeader)
        
        if initial_dt and initial_dt.isValid(): 
            self.calendar.setSelectedDate(initial_dt.date())
        body.addWidget(self.calendar)

        # Lógica de las ruedas de tiempo (TimeWheel)
        time_lay = QHBoxLayout()
        self.h_w = TimeWheel(24); self.m_w = TimeWheel(60); self.s_w = TimeWheel(60)
        t = initial_dt.time() if initial_dt else QTime(0,0,0)
        self.h_w.setCurrentRow(t.hour()); self.m_w.setCurrentRow(t.minute()); self.s_w.setCurrentRow(t.second())
        
        for w, l in zip([self.h_w, self.m_w, self.s_w], ["H", "M", "S"]):
            v = QVBoxLayout(); lbl = QLabel(l); lbl.setAlignment(Qt.AlignCenter); lbl.setStyleSheet("font-size: 9px; color: #999; border:none;")
            v.addWidget(lbl); v.addWidget(w); time_lay.addLayout(v)
        
        body.addLayout(time_lay)
        layout.addLayout(body)
        
        self.btn_apply = QPushButton("Aplicar")
        self.btn_apply.setFixedSize(70, 26)
        self.btn_apply.setCursor(Qt.PointingHandCursor)
        self.btn_apply.setStyleSheet("""
            QPushButton { background: #0078d7; color: white; border-radius: 3px; font-size: 11px; font-weight: bold; }
            QPushButton:hover { background: #005fa3; }
        """)
        self.btn_apply.clicked.connect(self.accept)
        
        bottom_lay = QHBoxLayout(); bottom_lay.addStretch(); bottom_lay.addWidget(self.btn_apply)
        layout.addLayout(bottom_lay)
        QVBoxLayout(self).addWidget(self.main_frame)

    def get_selected_dt(self):
        return QDateTime(self.calendar.selectedDate(), QTime(self.h_w.currentRow(), self.m_w.currentRow(), self.s_w.currentRow()))
    
class CustomDateTimePicker(QWidget):
    dateTimeChanged = Signal()

    def __init__(self):
        super().__init__()
        layout = QHBoxLayout(self); layout.setContentsMargins(0,0,0,0); layout.setSpacing(0)
        
        self.line_edit = QLineEdit()
        self.line_edit.setInputMask("99/99/9999 99:99:99") # Formato visual para el usuario
        self.line_edit.setFixedHeight(26)
        self.line_edit.setStyleSheet("QLineEdit { border: 1px solid #ccc; border-radius: 3px 0 0 3px; padding-left: 5px; color: #333; font-size: 11px; }")
        self.line_edit.textChanged.connect(self._check_validity)

        self.btn = QPushButton("📅")
        self.btn.setFixedSize(26, 26)
        self.btn.setStyleSheet("background: #f8f8f8; border: 1px solid #ccc; border-left: none; border-radius: 0 3px 3px 0; color: #666;")
        self.btn.clicked.connect(self.open_picker)
        
        layout.addWidget(self.line_edit); layout.addWidget(self.btn)

    def _check_validity(self):
        if self.dateTime().isValid():
            self.dateTimeChanged.emit()

    def open_picker(self):
        dt = self.dateTime()
        if not dt.isValid(): dt = QDateTime.currentDateTime()
        pop = DateTimePickerPopup(self, dt)
        pos = self.mapToGlobal(self.line_edit.rect().bottomLeft())
        pop.move(pos.x(), pos.y() + 1)
        if pop.exec_():
            self.setDateTime(pop.get_selected_dt())
            self.dateTimeChanged.emit()

    def setDateTime(self, dt):
        if dt.isValid():
            self.line_edit.setText(dt.toString("dd/MM/yyyy HH:mm:ss"))

    def dateTime(self):
        # Siempre parsear desde el formato del input mask
        return QDateTime.fromString(self.line_edit.text(), "dd/MM/yyyy HH:mm:ss")

class ExportarData():
    
    @staticmethod
    def validarExportarDataEquipos(idproyecto, nameproyecto, idzona, tipo, equipos, fechainicial=None, fechafinal=None):
        formato = "yyyy-MM-dd HH:mm:ss"
        formato_sql = formato
        # Crear el diálogo principal
        dialogo = QDialog()
        dialogo.setWindowTitle(f"Exportar data {tipo}")
        dialogo.setMinimumWidth(500)
        dialogo.setStyleSheet("background-color: white;")
        main_layout = QVBoxLayout(dialogo)
        main_layout.setContentsMargins(15, 15, 15, 15)
        # --- Diferencia de días
        header = QHBoxLayout()
        labeldias = QLabel("0")
        labeldias.setStyleSheet("color: #0078d7; font-weight: bold; font-size: 14px;")
        header.addWidget(labeldias)
        header.addWidget(QLabel("días seleccionados"))
        header.addStretch()
        main_layout.addLayout(header)
        # --- Selectores de fecha
        form_layout = QHBoxLayout()
        # Inicio
        v1 = QVBoxLayout()
        v1.addWidget(QLabel("DESDE"))
        dt_inicio = CustomDateTimePicker()
        v1.addWidget(dt_inicio)
        # Fin
        v2 = QVBoxLayout()
        v2.addWidget(QLabel("HASTA"))
        dt_final = CustomDateTimePicker()
        v2.addWidget(dt_final)
        form_layout.addLayout(v1)
        form_layout.addSpacing(10)
        form_layout.addLayout(v2)
        main_layout.addLayout(form_layout)
        # --- Botones de acción ---
        btn_layout = QHBoxLayout()
        botoncancelar = QPushButton("Cancelar")
        botoncancelar.clicked.connect(dialogo.reject)
        botonexportar = QPushButton("EXPORTAR")
        botonexportar.setFixedHeight(30)
        botonexportar.setStyleSheet("""
            QPushButton {
                background: #0078d7; color: white; font-weight: bold;
                border-radius: 4px; padding: 0 15px;
            }
            QPushButton:disabled {
                background: #f0f0f0; color: #ccc;
            }
        """)
        # Botón adicional para CSV (solo para Prismas)
        botonexportarcsv = QPushButton("CSV")
        botonexportarcsv.setFixedHeight(30)
        botonexportarcsv.setStyleSheet(botonexportar.styleSheet())
        botonexportarcsv.setVisible(tipo == "Prismas")
        btn_layout.addStretch()
        btn_layout.addWidget(botoncancelar)
        btn_layout.addWidget(botonexportarcsv)
        btn_layout.addWidget(botonexportar)
        main_layout.addLayout(btn_layout)
        # habilitar exportar solo si diferencia > 0
        def validar():
            ini = dt_inicio.dateTime()
            fin = dt_final.dateTime()
            if ini.isValid() and fin.isValid():
                labeldias.setText(str(ini.date().daysTo(fin.date())))
                botonexportar.setEnabled(ini.secsTo(fin) >= 60)
            else:
                botonexportar.setEnabled(False)

        # --- Funciones de exportación ---
        def exportarDataEquipo():
            time_inicio = dt_inicio.dateTime().toString(formato_sql)
            time_fin = dt_final.dateTime().toString(formato_sql)
            dialogo.accept()  # cerrar con éxito

            # Llamadas originales sin cambios
            if tipo == "Prismas":
                ExportarData.exportarExcelPrismas(idproyecto, nameproyecto, idzona, equipos, time_inicio, time_fin)
            elif tipo == "Inclinómetros":
                ExportarData.exportarZipInclinometros(idproyecto, nameproyecto, idzona, equipos)
            elif tipo == "Piezómetros Cuerda Vibrante":
                ExportarData.exportarExcelPiezometros("Automatizado", idproyecto, nameproyecto, idzona, equipos, time_inicio, time_fin)
            elif tipo == "Piezómetros Casagrande":
                ExportarData.exportarExcelPiezometros("Manual", idproyecto, nameproyecto, idzona, equipos, time_inicio, time_fin)
            elif tipo == "Pluviómetros":
                ExportarData.exportarExcelPluviometros(idproyecto, nameproyecto, idzona, equipos)
            elif tipo == "Cotas de Terreno":
                ExportarData.exportarExcelCotasTerreno(idproyecto, nameproyecto, idzona, equipos)
            elif tipo == "Celdas de Asentamiento":
                ExportarData.exportarExcelCeldasAsentamiento(idproyecto, nameproyecto, idzona, equipos, time_inicio, time_fin)
            elif tipo == "Acelerógrafos":
                ExportarData.exportarExcelAcelerografos(idproyecto, nameproyecto, idzona, equipos, time_inicio, time_fin)
            elif tipo == "TDR":
                ExportarData.exportarZipSondajesTDR(idproyecto, nameproyecto, idzona, equipos)

        def exportarDataEquiposCSV():
            time_inicio = dt_inicio.dateTime().toString(formato_sql)
            time_fin = dt_final.dateTime().toString(formato_sql)
            dialogo.accept()
            ExportarData.exportarExcelPrismasCSV(idproyecto, nameproyecto, idzona, equipos, time_inicio, time_fin)

        # Conectar señales
        dt_inicio.dateTimeChanged.connect(validar)
        dt_final.dateTimeChanged.connect(validar)
        botonexportar.clicked.connect(exportarDataEquipo)
        botonexportarcsv.clicked.connect(exportarDataEquiposCSV)

        # Cargar fechas iniciales
        def parsear_entrada(valor):
            if isinstance(valor, str):
                dt_parsed = QDateTime.fromString(valor, formato_sql)
                if not dt_parsed.isValid():
                    dt_parsed = QDateTime.fromString(valor, "dd/MM/yyyy HH:mm:ss")
                return dt_parsed
            elif isinstance(valor, datetime):
                return QDateTime(valor.year, valor.month, valor.day,
                                valor.hour, valor.minute, valor.second)
            return QDateTime()

        if fechainicial is not None:
            dt_ini = parsear_entrada(fechainicial)
            dt_fin = parsear_entrada(fechafinal)
            if dt_ini.isValid() and dt_fin.isValid():
                dt_inicio.setDateTime(dt_ini)
                dt_final.setDateTime(dt_fin)
            else:
                # Fallback a fecha/hora actual
                ahora = QDateTime.currentDateTime()
                dt_inicio.setDateTime(ahora)
                dt_final.setDateTime(ahora)
        else:
            ahora = QDateTime.currentDateTime()
            dt_inicio.setDateTime(ahora)
            dt_final.setDateTime(ahora)

        # Disparar validación inicial para actualizar el label y el botón
        validar()
        # Mostrar el diálogo (no necesitamos devolver valores, ya se procesa dentro)
        dialogo.exec()

    def exportarExcelPrismasCSV(proyectoid, proyectoname, idzona, prismasmarcados, fechaini, fechafin):
        encabezados = [
            'State', 'Point ID', 'Profile Name', 'Time', 'Hz', 'V', 'D [m]',
            'PPM Type', 'PPM', 'Pressure [mBar]', 'Av Temp [°C]', 'Add Const [m]',
            'Target Easting [m]', 'Target Northing [m]', 'Target Elevation [m]',
            'Reflector Height [m]', 'Instrument Height [m]', 'Station Easting [m]',
            'Station Northing [m]', 'Station Height [m]', 'Null Measurement [m]',
            'Short Time Diff [m]', 'Long Time Diff [m]', 'Vel Limit Diff [m]',
            'Horz Distance [m]', 'Difference Outlier Test [m]',
            'Longitudinal Displacement [m]',
            'Transverse Displacement [m]',
            'Height Displacement [m]', 'Point group'
        ]
        
        MAX_FILAS_POR_ARCHIVO = 1000000  # Máximo de filas por archivo CSV
        tabla = f"prismas{proyectoid}"
        nameprismas = [nameprisma for nameprisma, idinstrumento, tabla in prismasmarcados]
        # Obtener datos de la base de datos
        prismasdata = DatosController.ctrlObtenerDataExportarPrismas(tabla, nameprismas, fechaini, fechafin)
        if not prismasdata:
            return False
        
        total_filas = len(prismasdata)
        num_archivos = math.ceil(total_filas / MAX_FILAS_POR_ARCHIVO)
        # Permitir al usuario elegir la carpeta de destino
        carpeta_destino = QFileDialog.getExistingDirectory(
            None, 
            "Seleccionar carpeta para guardar archivos CSV",
            "",
            QFileDialog.ShowDirsOnly
        )
        
        if not carpeta_destino:
            return False
        
        try:
            archivos_creados = []
            for i in range(num_archivos):
                # Calcular rango de datos para este archivo
                inicio = i * MAX_FILAS_POR_ARCHIVO
                fin = min((i + 1) * MAX_FILAS_POR_ARCHIVO, total_filas)
                segmento = prismasdata[inicio:fin]
                
                # Generar nombre de archivo con sufijo numérico si hay múltiples archivos
                if num_archivos > 1:
                    nombre_archivo = f"datos_prismas_{idzona}_parte_{i+1}.csv"
                else:
                    nombre_archivo = f"datos_prismas_{idzona}.csv"
                
                # Crear ruta completa del archivo
                ruta_completa = os.path.join(carpeta_destino, nombre_archivo)
                
                # Escribir archivo CSV
                with open(ruta_completa, 'w', newline='', encoding='latin-1') as archivo_csv:
                    escritor = csv.writer(archivo_csv)
                    # Escribir encabezado en cada archivo
                    escritor.writerow(encabezados)
                    escritor.writerows(segmento)
                
                archivos_creados.append(ruta_completa)
            # Mostrar mensaje de éxito
            if num_archivos == 1:
                mensaje = f"Archivo CSV exportado exitosamente:\n{archivos_creados[0]}"
            else:
                mensaje = f"Se exportaron {num_archivos} archivos CSV en:\n{carpeta_destino}"
            
            mostrar_mensaje("Exportación exitosa", mensaje, "informacion")
            return True
            
        except Exception as e:
            return False
    
    def exportarExcelPrismas(proyectoid, proyectoname, idzona, prismasmarcados, fechaini, fechafin):
        config = SoftwareConfiguracion.obtenerDataSoftware()
        tipovelocidad = config[15]
        respuesta = EmpresaConfiguracion.obtenerDataEmpresa()
        logo = respuesta[4]
        libro = Workbook()
        cont = 0
        for nameprisma, idinstrumento, tabla in prismasmarcados:
            prismasdata = DatosController.ctrlObtenerPrismasDataExportar(proyectoid, tabla, idzona, nameprisma, tipovelocidad, fechaini, fechafin)
            if prismasdata:
                if cont == 0:
                    hoja = libro.active
                    hoja.title = nameprisma
                else:
                    hoja = libro.create_sheet(title=nameprisma)
                # Configurar la hoja (común para todas)
                ExportarData.configurarCabeceraHojaPrismas(hoja, proyectoname, nameprisma, logo)
                # Insertar datos en la tabla comenzando desde la fila 11
                for fila in prismasdata:
                    hoja.append(fila)
                cont += 1
        rutaexcel = resource_path("resources/workspace/dataequipos.xlsx")
        libro.save(rutaexcel)
        archivo_destino, _ = QFileDialog.getSaveFileName(None, "Guardar Excel en", "Prismas", "Archivos de Excel (*.xlsx);;Todos los archivos (*)")
        if archivo_destino:
            # Asegurarse de que el archivo tenga la extensión .xlsx
            if not archivo_destino.lower().endswith('.xlsx'):
                archivo_destino += '.xlsx'
            try:
                shutil.copy(rutaexcel, archivo_destino)
                mostrar_mensaje("Data Exportada", f"El Excel se ha guardado en: {archivo_destino}", "informacion")
            except Exception as e:  # Captura cualquier excepción
                mostrar_mensaje("Error al Exportar", f"No se pudo guardar el Excel: {str(e)}", "advertencia")
    
    def configurarCabeceraHojaPrismas(hoja, proyectoname, nameprisma, logo):
        # Definir colores, bordes, y otros atributos comunes
        color_fondo = PatternFill(start_color="3C3C3C", end_color="3C3C3C", fill_type="solid")
        borde_negro = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        # Ajustar el ancho de las columnas (A hasta K)
        for col in range(1, 12):
            hoja.column_dimensions[chr(64 + col)].width = 20
        # Ajustar el alto de las filas 1 a 4 para el logo
        for row in range(1, 5):
            hoja.row_dimensions[row].height = 25
        # ---- LOGO (A1:A4) ----
        if logo:
            imagen_stream = MetodosGenerales.convertirBlobImagen(logo)
            imagen = ExcelImage(imagen_stream)
        else:
            ui_file_path = resource_path("resources/logo.png")
            imagen = ExcelImage(ui_file_path)
        # Tamaño cuadrado 3.5 cm x 3.5 cm
        imagen.width = 132
        imagen.height = 132
        imagen.anchor = "A1"
        hoja.merge_cells("A1:A4")
        hoja.add_image(imagen)
        # ---- TÍTULO (B1:K4) ----
        hoja.merge_cells("B1:K4")
        celda_titulo = hoja["B1"]
        celda_titulo.value = f"MONITOREO DE HITOS TOPOGRÁFICOS - {proyectoname.upper()}"
        celda_titulo.font = Font(size=18, bold=True)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # ---- SUBTÍTULO (ahora en A6:K7) ----
        hoja.merge_cells("A6:K7")
        celda_subtitulo = hoja["A6"]
        celda_subtitulo.value = f"Hito Topográfico: {nameprisma}"
        celda_subtitulo.font = Font(size=16, bold=True, color="FFFFFF")
        celda_subtitulo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda_subtitulo.fill = color_fondo
        # ---- BORDES para las celdas fusionadas ----
        for rango in ["A1:A4", "B1:K4", "A6:K7",
                    "A9:C9", "D9:G9", "H9:I9", "J9:K9"]:
            for fila_celdas in hoja[rango]:
                for celda in fila_celdas:
                    celda.border = borde_negro
        # ---- CABECERAS DE GRUPO (ahora en fila 9) ----
        cabeceras = [
            ("A9:C9", "Datos"),
            ("D9:G9", "Coordenadas"),
            ("H9:I9", "Desplazamiento (m)"),
            ("J9:K9", "Velocidad (cm/día)"),
        ]
        for rango, texto in cabeceras:
            hoja.merge_cells(rango)
            celda = hoja[rango.split(":")[0]]
            celda.value = texto
            celda.font = Font(size=14, bold=True, color="FFFFFF")
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.fill = color_fondo
        # ---- ENCABEZADOS DETALLADOS (ahora en fila 10) ----
        encabezados = [
            "Hito", "Fecha", "Hora", "Este", "Norte", "Elevación",
            "Distancia", "Incremental", "Acumulado", "Incremental", "Acumulado"
        ]
        for col, encabezado in enumerate(encabezados, 1):
            celda = hoja.cell(row=10, column=col, value=encabezado)
            celda.font = Font(bold=True, color="FFFFFF")
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.fill = color_fondo
            celda.border = borde_negro
    
    def exportarZipInclinometros(idproyecto, nameproyecto, idzona, inclinometrosmarcados):
        zip_filename = resource_path("resources/workspace/zipequipos.zip")
        try:
            with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for nombreincli, idinstrumento, idinclino in inclinometrosmarcados:
                    infoinclino = DatosController.ctrlObtenerInfoExportarInclinometro(idzona, "INCLINOMETRO", idinstrumento)
                    if infoinclino:
                        for info in infoinclino:
                            idencabezado, tipoequipo = info[0], info[4]
                            dataencabeza = DatosController.ctrlObtenerDataExportarInclinometro(idproyecto, idencabezado)
                            if dataencabeza:
                                if tipoequipo == "RST":
                                    csv_file = io.StringIO()
                                    writer = csv.writer(csv_file)
                                    ExportarData.generarEncabezadoCSVInclinometro(writer, info, nameproyecto)
                                    for datafila in dataencabeza:
                                        writer.writerow(datafila)
                                    csv_content = csv_file.getvalue()
                                    csv_file.close()
                                    # Nombre del archivo CSV dentro del ZIP
                                    csv_name = f"{info[2]}_{info[5]}.csv".replace("/", "_").replace(":", "_")
                                    zipf.writestr(csv_name, csv_content)
                                else:
                                    gkn_file = io.StringIO()
                                    writer = csv.writer(gkn_file)
                                    ExportarData.generarEncabezadoGKNInclinometro(writer, info, nameproyecto)
                                    for datafila in dataencabeza:
                                        writer.writerow(datafila)
                                    gkn_content = gkn_file.getvalue()
                                    gkn_file.close()
                                    # Nombre del archivo con extensión .gkn
                                    gkn_name = f"{info[2]}_{info[5]}.gkn".replace("/", "_").replace(":", "_")
                                    zipf.writestr(gkn_name, gkn_content)
            # Guardar el archivo ZIP en una ubicación elegida por el usuario
            archivo_destino, _ = QFileDialog.getSaveFileName(
                None, "Guardar ZIP en", "Inclinometros", "Archivos ZIP (*.zip);;Todos los archivos (*)"
            )
            if archivo_destino:
                if not archivo_destino.lower().endswith('.zip'):
                    archivo_destino += '.zip'
                shutil.copy(zip_filename, archivo_destino)
                mostrar_mensaje("Data Exportada", f"El ZIP se ha guardado en: {archivo_destino}", "informacion")
        except Exception as e:
            mostrar_mensaje("Error al Exportar", f"No se pudo guardar el ZIP: {str(e)}", "advertencia")
    
    def generarEncabezadoCSVInclinometro(writer, infoinclino, nombreproyecto):
        writer.writerow(["RST Digital Inclinometer Data"])
        writer.writerow(["File Version", ""])
        writer.writerow(["File Type", "Digital Inclinometer"])
        writer.writerow(["Site", nombreproyecto])
        writer.writerow(["Zone", infoinclino[1]])
        writer.writerow(["Borehole", infoinclino[2]])
        writer.writerow(["Probe Serial", infoinclino[3]])
        fechaorig, hora = infoinclino[5].split(" ")
        fecha_obj = datetime.strptime(fechaorig, "%Y-%m-%d")
        fecha = fecha_obj.strftime("%m/%d/%Y")
        writer.writerow(["Reading Date(m/d/y)", fecha, hora])
        writer.writerow(["Depth", infoinclino[6]])
        writer.writerow(["Interval", "0.5"])
        writer.writerow(["Depth Units", "m"])
        writer.writerow(["Reading Units", ""])
        writer.writerow(["Offset Correction", ""])
        writer.writerow(["East", infoinclino[7]])
        writer.writerow(["North", infoinclino[8]])
        writer.writerow(["Level", infoinclino[9]])
        writer.writerow([])
        writer.writerow(["Depth", "Face A+", "Face A-", "Face B+", "Face B-"])
        writer.writerow([])
    
    def generarEncabezadoGKNInclinometro(writer, infoinclino, nombreproyecto):
        writer.writerow(["GEOKON Inclinometer Data"])
        writer.writerow(["GKN FORMAT", ""])
        writer.writerow([f"PROJECT  :{nombreproyecto}"])
        writer.writerow([f"ZONE     :{infoinclino[1]}"])
        writer.writerow([f"HOLE NO. :{infoinclino[2]}"])
        fechaorig, hora = infoinclino[5].split(" ")
        fecha_obj = datetime.strptime(fechaorig, "%Y-%m-%d")
        fecha = fecha_obj.strftime("%m/%d/%y")
        writer.writerow([f"DATE     :{fecha}"])
        writer.writerow([f"TIME     :{hora}"])
        writer.writerow([f"PROBE NO.:{infoinclino[3]}"])
        writer.writerow([f"COORD.   :({infoinclino[7]}, {infoinclino[8]}, {infoinclino[9]})"])
        writer.writerow([f"#READINGS:"])
        writer.writerow(["FLEVEL", "A+", "A-", "B+", "B-"])
    
    def exportarExcelPiezometros(tipo, idproyecto, nameproyecto, idzona, piezometrosmarcados, fechaini, fechafin):
        respuesta = EmpresaConfiguracion.obtenerDataEmpresa()
        logo = respuesta[4]
        libro = Workbook()
        cont = 0
        for namepiezo, idinstrumento, idpiezo in piezometrosmarcados:
            if tipo == "Automatizado":
                tipoequipo = "PIEZOMETROCUERDA"
            else:
                tipoequipo = "PIEZOMETROMANUAL"
            infopiezo = DatosController.ctrlTraerInfoPiezometro(tipo, idzona, tipoequipo, idinstrumento)
            datospiezo = DatosController.ctrlListarDataPiezometrosProyecto(tipo, idproyecto, idzona, idinstrumento, fechaini, fechafin)
            if infopiezo and datospiezo:
                if cont == 0:
                    hoja = libro.active
                    hoja.title = namepiezo
                else:
                    hoja = libro.create_sheet(title=namepiezo)
                # Configurar la hoja (común para todas)
                if tipo == "Automatizado":
                    ExportarData.configurarCabeceraHojaPiezometroCuerda(hoja, infopiezo, nameproyecto, namepiezo, logo)
                else:
                    ExportarData.configurarCabeceraHojaPiezometroManual(hoja, infopiezo, nameproyecto, namepiezo, logo)
                # Insertar datos en la tabla comenzando desde la fila 12
                for fila in datospiezo:
                    hoja.append(list(fila))
                cont += 1
        rutaexcel = resource_path("resources/workspace/dataequipos.xlsx")
        libro.save(rutaexcel)
        archivo_destino, _ = QFileDialog.getSaveFileName(None, "Guardar Excel en", "Piezometros", "Archivos de Excel (*.xlsx);;Todos los archivos (*)")
        if archivo_destino:
            # Asegurarse de que el archivo tenga la extensión .xlsx
            if not archivo_destino.lower().endswith('.xlsx'):
                archivo_destino += '.xlsx'
            try:
                shutil.copy(rutaexcel, archivo_destino)
                mostrar_mensaje("Data Exportada", f"El Excel se ha guardado en: {archivo_destino}", "informacion")
            except Exception as e:  # Captura cualquier excepción
                mostrar_mensaje("Error al Exportar", f"No se pudo guardar el Excel: {str(e)}", "advertencia")
    
    def configurarCabeceraHojaPiezometroCuerda(hoja, infopiezo, proyectoname, namepiezo, logo):
        # Definir colores, bordes, y otros atributos comunes
        color_fondo = PatternFill(start_color="3C3C3C", end_color="3C3C3C", fill_type="solid")
        borde_negro = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        # Ajustar el ancho de las columnas para mejor presentación
        for col in range(1, 10):  # 9 columnas de ancho
            hoja.column_dimensions[chr(64 + col)].width = 20
        # Ajustar el alto de las filas de 1 a 5 para el área de la imagen
        for row in range(1, 5):
            hoja.row_dimensions[row].height = 25
        # Ruta de la imagen
        if logo:
            imagen_stream = MetodosGenerales.convertirBlobImagen(logo)
            imagen = ExcelImage(imagen_stream)
        else:
            ui_file_path = resource_path("resources/logo.png")
            imagen = ExcelImage(ui_file_path)
        imagen.width = 132
        imagen.height = 132
        imagen.anchor = "A1"
        hoja.merge_cells("A1:A4")
        hoja.add_image(imagen)
        # Agregar el título
        hoja.merge_cells("B1:I4")
        celda_titulo = hoja["B1"]
        celda_titulo.value = f"MONITOREO DE PIEZÓMETROS CUERDA VIBRANTE - {proyectoname.upper()}"
        celda_titulo.font = Font(size=18, bold=True)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Agregar el subtítuloS
        cabeceraGeneral = [("A6:C7", "DATOS DEL SENSOR"), ("D6:F7", "DATOS DE INSTALACIÓN"), ("G6:I7", "DATOS DE CALIBRACIÓN")]
        for rango, texto in cabeceraGeneral:
            hoja.merge_cells(rango)
            celda = hoja[rango.split(":")[0]]
            celda.value = texto
            celda.font = Font(size=14, bold=True, color="FFFFFF")
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.fill = color_fondo
        # Definir datos generales
        datosceldas = [("A8", "Nombre:"), ("A9", "Componente:"), ("A10", "Marca/Modelo:"), ("A11", "Serie:"),
                    ("F8", infopiezo[7]), ("F9", infopiezo[8]), ("F10", infopiezo[5]), ("F11", infopiezo[6]),
                    ("G8", "C.F.:"), ("G9", "T.K.:"), ("G10", "Inclinación:"), ("G11", "Azimuth:")]
        datoscombinados = [("B8:C8", namepiezo), ("B9:C9", infopiezo[24]), ("B10:C10", ""), ("B11:C11", infopiezo[4]),
                        ("D8:E8", "Cota Instalación (m.s.n.m):"), ("D9:E9", "Cota Fundación (m.s.n.m):"),
                        ("D10:E10", "Coordenada Este:"), ("D11:E11", "Coordenada Norte:"), ("H8:I8", infopiezo[11]),
                        ("H9:I9", infopiezo[12]), ("H10:I10", infopiezo[9]), ("H11:I11", infopiezo[10])]
        for rango, texto in datosceldas:
            celda = hoja[rango]
            celda.value = texto
            if rango.startswith("F"):
                celda.alignment = Alignment(horizontal="center", vertical="center")
            else:
                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = color_fondo
        for rango, texto in datoscombinados:
            hoja.merge_cells(rango)
            celda = hoja[rango.split(":")[0]]
            celda.value = texto
            if rango.startswith("D") or rango.startswith("E"):
                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = color_fondo
            else:
                celda.alignment = Alignment(horizontal="center", vertical="center")
        # Agregar bordes a las celdas fusionadas
        rangosceldas = ["A1:A4", "B1:I4", "A6:C7", "D6:F7", "G6:I7", "A8:I11"]
        for rango in rangosceldas:
            for fila_celdas in hoja[rango]:
                for celda in fila_celdas:
                    celda.border = borde_negro
        # Fila de encabezados detallados
        encabezados = ["Fecha", "Hora", f"Frecuencia ({infopiezo[16]})", "Temperatura (°C)", f"Presión ({infopiezo[16]})", "mca (m)", "Cota Piezométrica", "Observación", "Cota Superficie"]
        for col, encabezado in enumerate(encabezados, 1):
            hoja[chr(64 + col) + "13"].value = encabezado
            celda = hoja[chr(64 + col) + "13"]
            celda.font = Font(bold=True, color="FFFFFF")
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.fill = color_fondo
            celda.border = borde_negro
    
    def configurarCabeceraHojaPiezometroManual(hoja, infopiezo, proyectoname, namepiezo, logo):
        # Definir colores, bordes, y otros atributos comunes
        color_fondo = PatternFill(start_color="3C3C3C", end_color="3C3C3C", fill_type="solid")
        borde_negro = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        # Ajustar el ancho de las columnas para mejor presentación
        for col in range(1, 11):  # 10 columnas de ancho
            hoja.column_dimensions[chr(64 + col)].width = 20
        # Ajustar el alto de las filas de 1 a 5 para el área de la imagen
        for row in range(1, 5):
            hoja.row_dimensions[row].height = 25
        # Ruta de la imagen
        if logo:
            imagen_stream = MetodosGenerales.convertirBlobImagen(logo)
            imagen = ExcelImage(imagen_stream)
        else:
            ui_file_path = resource_path("resources/logo.png")
            imagen = ExcelImage(ui_file_path)
        imagen.width = 132
        imagen.height = 132
        imagen.anchor = "A1"
        hoja.merge_cells("A1:A4")
        hoja.add_image(imagen)
        # Agregar el título
        hoja.merge_cells("B1:H4")
        celda_titulo = hoja["B1"]
        celda_titulo.value = f"MONITOREO DE PIEZÓMETROS CASAGRANDE - {proyectoname.upper()}"
        celda_titulo.font = Font(size=18, bold=True)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Agregar el subtítuloS
        cabeceraGeneral = [("A7:C8", "DATOS DEL EQUIPO"), ("D7:H8", "DATOS DE INSTALACIÓN")]
        for rango, texto in cabeceraGeneral:
            hoja.merge_cells(rango)
            celda = hoja[rango.split(":")[0]]
            celda.value = texto
            celda.font = Font(size=14, bold=True, color="FFFFFF")
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.fill = color_fondo
        # Definir datos generales
        datosceldas = [("A9", "Nombre:"), ("A10", "Código:"), ("A11", "Tipo:"), ("A12", "Ubicación:"),
                    ("F9", infopiezo[6]), ("F10", infopiezo[7]), ("F11", infopiezo[5]), ("F12", infopiezo[4]),
                    ("G9", "Inclinación:"), ("G10", "Azimuth:"), ("G11", "Stick Up (m):"), ("G12", "Comentario:"),
                    ("H9", infopiezo[8]), ("H10", infopiezo[9]), ("H11", infopiezo[10]), ("H12", infopiezo[11])]
        datoscombinados = [("B9:C9", namepiezo), ("B10:C10", infopiezo[3]), ("B11:C11", ""), ("B12:C12", infopiezo[14]),
                        ("D9:E9", "Cota Fondo Pozo (m.s.n.m):"), ("D10:E10", "Cota Fundación (m.s.n.m):"),
                        ("D11:E11", "Coordenada Este:"), ("D12:E12", "Coordenada Norte:")]
        for rango, texto in datosceldas:
            celda = hoja[rango]
            celda.value = texto
            if rango.startswith("F") or rango.startswith("H"):
                celda.alignment = Alignment(horizontal="center", vertical="center")
            else:
                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = color_fondo
        for rango, texto in datoscombinados:
            hoja.merge_cells(rango)
            celda = hoja[rango.split(":")[0]]
            celda.value = texto
            if rango.startswith("D") or rango.startswith("E"):
                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = color_fondo
            else:
                celda.alignment = Alignment(horizontal="center", vertical="center")
        # Agregar bordes a las celdas fusionadas
        rangosceldas = ["A1:D4", "E1:H4", "A7:D8", "E7:H8", "A9:H12"]
        for rango in rangosceldas:
            for fila_celdas in hoja[rango]:
                for celda in fila_celdas:
                    celda.border = borde_negro
        # Fila de encabezados detallados
        encabezados = ["Fecha", "Hora", "Niv. Piezométrico (m)", "Profundidad (m)", "Elevación (m.s.n.m)", "Cota Piezométrica", "Nivel Vertical (m)", "Observación"]
        for col, encabezado in enumerate(encabezados, 1):
            hoja[chr(64 + col) + "14"].value = encabezado
            celda = hoja[chr(64 + col) + "14"]
            celda.font = Font(bold=True, color="FFFFFF")
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.fill = color_fondo
            celda.border = borde_negro
    
    def exportarExcelPluviometros(idproyecto, nameproyecto, idzona, pluviometrosmarcados):
        respuesta = EmpresaConfiguracion.obtenerDataEmpresa()
        logo = respuesta[4]
        libro = Workbook()
        cont = 0
        for namepluvio, idinstrumento, idpluvio in pluviometrosmarcados:
            infopluvio = DatosController.ctrlTraerInfoPluviometro(idzona, "PLUVIOMETRO", idinstrumento)
            datospluvio = DatosController.ctrlListarDataPluviometro(idproyecto, idzona, idinstrumento)
            if infopluvio and datospluvio:
                if cont == 0:
                    hoja = libro.active
                    hoja.title = namepluvio
                else:
                    hoja = libro.create_sheet(title=namepluvio)
                # Configurar la hoja (común para todas)
                ExportarData.configurarCabeceraHojaPluviometro(hoja, infopluvio, nameproyecto, namepluvio, logo)
                # Insertar datos en la tabla comenzando desde la fila 19
                for fila in datospluvio:
                    hoja.append(list(fila))
                cont += 1
        rutaexcel = resource_path("resources/workspace/dataequipos.xlsx")
        libro.save(rutaexcel)
        archivo_destino, _ = QFileDialog.getSaveFileName(None, "Guardar Excel en", "Pluviometros", "Archivos de Excel (*.xlsx);;Todos los archivos (*)")
        if archivo_destino:
            # Asegurarse de que el archivo tenga la extensión .xlsx
            if not archivo_destino.lower().endswith('.xlsx'):
                archivo_destino += '.xlsx'
            try:
                shutil.copy(rutaexcel, archivo_destino)
                mostrar_mensaje("Data Exportada", f"El Excel se ha guardado en: {archivo_destino}", "informacion")
            except Exception as e:  # Captura cualquier excepción
                mostrar_mensaje("Error al Exportar", f"No se pudo guardar el Excel: {str(e)}", "advertencia")
    
    def configurarCabeceraHojaPluviometro(hoja, infopluvio, proyectoname, namepluvio, logo):
        # Definir colores, bordes, y otros atributos comunes
        color_fondo = PatternFill(start_color="3C3C3C", end_color="3C3C3C", fill_type="solid")
        borde_negro = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        # Ajustar el ancho de las columnas para mejor presentación
        for col in range(1, 5):  # 4 columnas de ancho
            hoja.column_dimensions[chr(64 + col)].width = 20
        # Ajustar el alto de las filas de 1 a 5 para el área de la imagen
        for row in range(1, 5):
            hoja.row_dimensions[row].height = 25
        # Ruta de la imagen
        if logo:
            imagen_stream = MetodosGenerales.convertirBlobImagen(logo)
            imagen = ExcelImage(imagen_stream)
        else:
            ui_file_path = resource_path("resources/logo.png")
            imagen = ExcelImage(ui_file_path)
        imagen.width = 132
        imagen.height = 132
        imagen.anchor = "A1"
        hoja.merge_cells("A1:A4")
        hoja.add_image(imagen)
        # Agregar el título
        hoja.merge_cells("B1:D4")
        celda_titulo = hoja["B1"]
        celda_titulo.value = f"MONITOREO DE PLUVIÓMETROS - {proyectoname.upper()}"
        celda_titulo.font = Font(size=18, bold=True)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Agregar el subtítuloS
        rango, texto = "A6:D7", "DATOS DEL EQUIPO"
        hoja.merge_cells(rango)
        celda = hoja[rango.split(":")[0]]
        celda.value = texto
        celda.font = Font(size=14, bold=True, color="FFFFFF")
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.fill = color_fondo
        # Definir datos generales
        datosceldas = [("A8", "Nombre:"), ("A9", "Código:"), ("A10", "Este:"),
                    ("C8", "Norte:"), ("C9", "Elevación:"), ("C10", "Comentario:")]
        datoscombinados = [("B8", namepluvio), ("B9", infopluvio[3]), ("B10", infopluvio[4]),
                        ("D8", infopluvio[5]), ("D9", infopluvio[6]), ("D10", infopluvio[7])]
        for rango, texto in datosceldas:
            celda = hoja[rango]
            celda.value = texto
            celda.border = borde_negro
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = color_fondo
        for rango, texto in datoscombinados:
            hoja.merge_cells(rango)
            celda = hoja[rango.split(":")[0]]
            celda.value = texto
            celda.border = borde_negro
            celda.alignment = Alignment(horizontal="left")
        # Agregar bordes a las celdas fusionadas
        for rango in ["A1:D4", "A6:D7"]:
            for fila_celdas in hoja[rango]:
                for celda in fila_celdas:
                    celda.border = borde_negro
        # Fila de encabezados detallados
        encabezados = ["Fecha", "Hora", "Precipitación (mm)", "Observación"]
        for col, encabezado in enumerate(encabezados, 1):
            hoja[chr(64 + col) + "12"].value = encabezado
            celda = hoja[chr(64 + col) + "12"]
            celda.font = Font(bold=True, color="FFFFFF")
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.fill = color_fondo
            celda.border = borde_negro
    
    def exportarExcelCeldasAsentamiento(idproyecto, nameproyecto, idzona, celdasmarcadas, fechaini, fechafin):
        respuesta = EmpresaConfiguracion.obtenerDataEmpresa()
        logo = respuesta[4]
        libro = Workbook()
        cont = 0
        for namecelda, idinstrumento, idcelda in celdasmarcadas:
            infocelda = DatosController.ctrlTraerInfoCeldaAsentamiento(idzona, "CELDA", idinstrumento)
            datoscelda = DatosController.ctrlListarDataCeldaAsentamiento(idproyecto, idzona, idinstrumento, fechaini, fechafin)
            if infocelda and datoscelda:
                if cont == 0:
                    hoja = libro.active
                    hoja.title = namecelda
                else:
                    hoja = libro.create_sheet(title=namecelda)
                # Configurar la hoja (común para todas)
                ExportarData.configurarCabeceraHojaCeldaAsentamiento(hoja, infocelda, nameproyecto, namecelda, logo)
                # Insertar datos en la tabla comenzando desde la fila 12
                for fila in datoscelda:
                    hoja.append(list(fila))
                cont += 1
        rutaexcel = resource_path("resources/workspace/dataequipos.xlsx")
        libro.save(rutaexcel)
        archivo_destino, _ = QFileDialog.getSaveFileName(None, "Guardar Excel en", "Celdas", "Archivos de Excel (*.xlsx);;Todos los archivos (*)")
        if archivo_destino:
            # Asegurarse de que el archivo tenga la extensión .xlsx
            if not archivo_destino.lower().endswith('.xlsx'):
                archivo_destino += '.xlsx'
            try:
                shutil.copy(rutaexcel, archivo_destino)
                mostrar_mensaje("Data Exportada", f"El Excel se ha guardado en: {archivo_destino}", "informacion")
            except Exception as e:  # Captura cualquier excepción
                mostrar_mensaje("Error al Exportar", f"No se pudo guardar el Excel: {str(e)}", "advertencia")
    
    def configurarCabeceraHojaCeldaAsentamiento(hoja, infocelda, proyectoname, namecelda, logo):
        # Definir colores, bordes, y otros atributos comunes
        color_fondo = PatternFill(start_color="3C3C3C", end_color="3C3C3C", fill_type="solid")
        borde_negro = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        # Ajustar el ancho de las columnas para mejor presentación
        for col in range(1, 11):  # 10 columnas de ancho
            hoja.column_dimensions[chr(64 + col)].width = 20
        # Ajustar el alto de las filas de 1 a 5 para el área de la imagen
        for row in range(1, 5):
            hoja.row_dimensions[row].height = 25
        # Ruta de la imagen
        if logo:
            imagen_stream = MetodosGenerales.convertirBlobImagen(logo)
            imagen = ExcelImage(imagen_stream)
        else:
            ui_file_path = resource_path("resources/logo.png")
            imagen = ExcelImage(ui_file_path)
        imagen.width = 132
        imagen.height = 132
        imagen.anchor = "A1"
        hoja.merge_cells("A1:A4")
        hoja.add_image(imagen)
        # Agregar el título
        hoja.merge_cells("B1:I4")
        celda_titulo = hoja["B1"]
        celda_titulo.value = f"MONITOREO DE CELDAS DE ASENTAMIENTO - {proyectoname.upper()}"
        celda_titulo.font = Font(size=18, bold=True)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Agregar el subtítuloS
        cabeceraGeneral = [("A7:C8", "DATOS DEL SENSOR"), ("D7:G8", "DATOS DE INSTALACIÓN"), ("H7:I8", "DATOS DE CALIBRACIÓN")]
        for rango, texto in cabeceraGeneral:
            hoja.merge_cells(rango)
            celda = hoja[rango.split(":")[0]]
            celda.value = texto
            celda.font = Font(size=14, bold=True, color="FFFFFF")
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.fill = color_fondo
        # Definir datos generales
        datosceldas = [("A9", "Nombre:"), ("A10", "Marca:"), ("A11", "Modelo:"), ("A12", "Serie:"),
                    ("H9", "C.F.:"), ("H10", "T.K.:"), ("H11", "Temperatura:"), ("H12", "Rango:"), ("I9", infocelda[13]),
                    ("I10", infocelda[14]), ("I11", infocelda[12]), ("I12", infocelda[6])]
        datoscombinados = [("B9:C9", namecelda), ("B10:C10", infocelda[3]), ("B11:C11", infocelda[4]), ("B12:C12", infocelda[5]),
                        ("D9:E9", "Cota Instalación (m.s.n.m):"), ("D10:E10", "Cota Fundación (m.s.n.m):"),
                        ("D11:E11", "Coordenada Este:"), ("D12:E12", "Coordenada Norte:"),
                        ("F9:G9", infocelda[9]), ("F10:G10", infocelda[10]), ("F11:G11", infocelda[7]), ("F12:G12", infocelda[8])]
        for rango, texto in datosceldas:
            celda = hoja[rango]
            celda.value = texto
            if rango.startswith("F") or rango.startswith("I"):
                celda.alignment = Alignment(horizontal="center", vertical="center")
            else:
                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = color_fondo
        for rango, texto in datoscombinados:
            hoja.merge_cells(rango)
            celda = hoja[rango.split(":")[0]]
            celda.value = texto
            if rango.startswith("D"):
                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = color_fondo
            else:
                celda.alignment = Alignment(horizontal="center", vertical="center")
        # Agregar bordes a las celdas fusionadas
        rangosceldas = ["A1:A4", "B1:I4", "A7:C8", "D7:G8", "H7:I8", "A9:I12"]
        for rango in rangosceldas:
            for fila_celdas in hoja[rango]:
                for celda in fila_celdas:
                    celda.border = borde_negro
        # Fila de encabezados detallados
        encabezados = ["Fecha", "Hora", "Frecuencia (Digits)", "Frecuencia (Hz)", "Temperatura (°C)", "Desplazamiento (m)", "Cota (m.s.n.m)", "Superficie (m.s.n.m)", "Observación"]
        for col, encabezado in enumerate(encabezados, 1):
            hoja[chr(64 + col) + "14"].value = encabezado
            celda = hoja[chr(64 + col) + "14"]
            celda.font = Font(bold=True, color="FFFFFF")
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.fill = color_fondo
            celda.border = borde_negro
    
    def exportarExcelAcelerografos(idproyecto, nameproyecto, idzona, acelerografosmarcados, fechaini, fechafin):
        respuesta = EmpresaConfiguracion.obtenerDataEmpresa()
        logo = respuesta[4]
        libro = Workbook()
        cont = 0
        for nameacelero, idinstrumento, idacelero in acelerografosmarcados:
            infoacelero = DatosController.ctrlTraerInfoAcelerografo(idzona, "ACELEROGRAFO", idinstrumento)
            datosacelero = DatosController.ctrlListarDataAcelerografo(idproyecto, idzona, idinstrumento, fechaini, fechafin)
            if infoacelero and datosacelero:
                if cont == 0:
                    hoja = libro.active
                    hoja.title = nameacelero
                else:
                    hoja = libro.create_sheet(title=nameacelero)
                # Configurar la hoja (común para todas)
                ExportarData.configurarCabeceraHojaAcelerografo(hoja, infoacelero, nameproyecto, nameacelero, logo)
                # Insertar datos en la tabla comenzando desde la fila 19
                for fila in datosacelero:
                    hoja.append(list(fila))
                cont += 1
        rutaexcel = resource_path("resources/workspace/dataequipos.xlsx")
        libro.save(rutaexcel)
        archivo_destino, _ = QFileDialog.getSaveFileName(None, "Guardar Excel en", "Acelerografos", "Archivos de Excel (*.xlsx);;Todos los archivos (*)")
        if archivo_destino:
            # Asegurarse de que el archivo tenga la extensión .xlsx
            if not archivo_destino.lower().endswith('.xlsx'):
                archivo_destino += '.xlsx'
            try:
                shutil.copy(rutaexcel, archivo_destino)
                mostrar_mensaje("Data Exportada", f"El Excel se ha guardado en: {archivo_destino}", "informacion")
            except Exception as e:  # Captura cualquier excepción
                mostrar_mensaje("Error al Exportar", f"No se pudo guardar el Excel: {str(e)}", "advertencia")
    
    def configurarCabeceraHojaAcelerografo(hoja, infoacelero, proyectoname, namepluvio, logo):
        # Definir colores, bordes, y otros atributos comunes
        color_fondo = PatternFill(start_color="3C3C3C", end_color="3C3C3C", fill_type="solid")
        borde_negro = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        # Ajustar el ancho de las columnas para mejor presentación
        for col in range(1, 6):  # 5 columnas de ancho
            hoja.column_dimensions[chr(64 + col)].width = 20
        # Ajustar el alto de las filas de 1 a 5 para el área de la imagen
        for row in range(1, 5):
            hoja.row_dimensions[row].height = 25
        # Ruta de la imagen
        if logo:
            imagen_stream = MetodosGenerales.convertirBlobImagen(logo)
            imagen = ExcelImage(imagen_stream)
        else:
            ui_file_path = resource_path("resources/logo.png")
            imagen = ExcelImage(ui_file_path)
        imagen.width = 132
        imagen.height = 132
        imagen.anchor = "A1"
        hoja.merge_cells("A1:A4")
        hoja.add_image(imagen)
        # Agregar el título
        hoja.merge_cells("B1:E4")
        celda_titulo = hoja["B1"]
        celda_titulo.value = f"MONITOREO DE ACELERÓGRAFOS - {proyectoname.upper()}"
        celda_titulo.font = Font(size=18, bold=True)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Agregar el subtítuloS
        rango, texto = "A6:E7", "DATOS DEL EQUIPO"
        hoja.merge_cells(rango)
        celda = hoja[rango.split(":")[0]]
        celda.value = texto
        celda.font = Font(size=14, bold=True, color="FFFFFF")
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.fill = color_fondo
        # Definir datos generales
        datosceldas = [("A8", "Nombre:"), ("A9", "Este:"), ("C8", "Norte:"), ("C9", "Elevación:"),
                       ("B8", namepluvio), ("B9", infoacelero[3])]
        datoscombinados = [("D8:E8", infoacelero[4]), ("D9:E9", infoacelero[5])]
        for rango, texto in datosceldas:
            celda = hoja[rango]
            celda.value = texto
            celda.border = borde_negro
            if rango.startswith("B"):
                celda.alignment = Alignment(horizontal="left")
            else:
                celda.alignment = Alignment(horizontal="center", vertical="center")
                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = color_fondo
        for rango, texto in datoscombinados:
            hoja.merge_cells(rango)
            celda = hoja[rango.split(":")[0]]
            celda.value = texto
            celda.alignment = Alignment(horizontal="center", vertical="center")
        # Agregar bordes a las celdas fusionadas
        for rango in ["A1:A4", "B1:E4", "A6:E7", "D8:E8", "D9:E9"]:
            for fila_celdas in hoja[rango]:
                for celda in fila_celdas:
                    celda.border = borde_negro
        # Fila de encabezados detallados
        encabezados = ["Fecha", "Hora", "Magnitud", "Distancia (Km)", "Observación"]
        for col, encabezado in enumerate(encabezados, 1):
            hoja[chr(64 + col) + "11"].value = encabezado
            celda = hoja[chr(64 + col) + "11"]
            celda.font = Font(bold=True, color="FFFFFF")
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.fill = color_fondo
            celda.border = borde_negro
    
    def exportarZipSondajesTDR(idproyecto, nameproyecto, idzona, sondajesmarcados):
        zip_filename = resource_path("resources/workspace/zipequipos.zip")
        try:
            respuesta = EmpresaConfiguracion.obtenerDataEmpresa()
            logo = respuesta[4]
            with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for nombretdr, idinstrumento, idtdr in sondajesmarcados:
                    infotdr = DatosController.ctrlObtenerInfoExportarSondajetdr(idproyecto, idzona, idinstrumento)
                    if infotdr:
                        workbook = Workbook()
                        first_sheet = True
                        for info in infotdr:
                            fecha = info[0]
                            dataencabeza = DatosController.ctrlObtenerDataExportarSondajetdr(idproyecto, idtdr, fecha)
                            if dataencabeza:
                                # Crear o seleccionar la hoja
                                nombretitle = f"{nombretdr}_{fecha}".replace("/", "_").replace(":", "_")
                                if first_sheet:
                                    sheet = workbook.active
                                    sheet.title = nombretitle
                                    first_sheet = False
                                else:
                                    sheet = workbook.create_sheet(title=nombretitle)
                                # Configurar cabecera y datos
                                ExportarData.configurarCabeceraHojaSondajetdr(sheet, info, nameproyecto, nombretdr, logo)
                                for fila in dataencabeza:
                                    sheet.append(list(fila))
                        # Guardar en memoria
                        excel_stream = io.BytesIO()
                        workbook.save(excel_stream)
                        excel_stream.seek(0)
                        # Nombre del archivo Excel dentro del ZIP
                        excel_name = f"{nombretdr}.xlsx"
                        zipf.writestr(excel_name, excel_stream.getvalue())
            # Guardar ZIP en una ubicación elegida por el usuario
            archivo_destino, _ = QFileDialog.getSaveFileName(None, "Guardar ZIP en", "SondajesTDR", "Archivos ZIP (*.zip);;Todos los archivos (*)")
            if archivo_destino:
                if not archivo_destino.lower().endswith('.zip'):
                    archivo_destino += '.zip'
                shutil.copy(zip_filename, archivo_destino)
                mostrar_mensaje("Data Exportada", f"El ZIP se ha guardado en: {archivo_destino}", "informacion")
        except Exception as e:
            mostrar_mensaje("Error al Exportar", f"No se pudo guardar el ZIP: {str(e)}", "advertencia")
    
    def configurarCabeceraHojaSondajetdr(hoja, infotdr, proyectoname, nametdr, logo):
        # Definir colores, bordes, y otros atributos comunes
        color_fondo = PatternFill(start_color="3C3C3C", end_color="3C3C3C", fill_type="solid")
        borde_negro = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        # Ajustar el ancho de las columnas para mejor presentación
        for col in range(1, 6):  # 5 columnas de ancho
            hoja.column_dimensions[chr(64 + col)].width = 20
        # Ajustar el alto de las filas de 1 a 5 para el área de la imagen
        for row in range(1, 5):
            hoja.row_dimensions[row].height = 25
        # Ruta de la imagen
        if logo:
            imagen_stream = MetodosGenerales.convertirBlobImagen(logo)
            imagen = ExcelImage(imagen_stream)
        else:
            ui_file_path = resource_path("resources/logo.png")
            imagen = ExcelImage(ui_file_path)
        imagen.width = 132
        imagen.height = 132
        imagen.anchor = "A1"
        hoja.merge_cells("A1:A4")
        hoja.add_image(imagen)
        # Agregar el título
        hoja.merge_cells("B1:E4")
        celda_titulo = hoja["B1"]
        celda_titulo.value = f"MONITOREO DE EQUIPOS TDR - {proyectoname.upper()}"
        celda_titulo.font = Font(size=18, bold=True)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Agregar el subtítuloS
        rango, texto = "A6:E7", "DATOS DEL EQUIPO"
        hoja.merge_cells(rango)
        celda = hoja[rango.split(":")[0]]
        celda.value = texto
        celda.font = Font(size=14, bold=True, color="FFFFFF")
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.fill = color_fondo
        # Definir datos generales
        datosceldas = [("A8", "Nombre:"), ("A9", "Este:"), ("A10", "Norte:"), ("A11", "Elevación:"),
                       ("B8", nametdr), ("B9", infotdr[2]), ("B10", infotdr[3]), ("B11", infotdr[4]),
                       ("C8", "Profundidad:"), ("C9", "Inclinación:"), ("C10", "Azimuth:"), ("C11", "Comentario:")]
        datoscombinados = [("D8:E8", infotdr[5]), ("D9:E9", infotdr[6]),
                           ("D10:E10", infotdr[7]), ("D11:E11", infotdr[8])]
        for rango, texto in datosceldas:
            celda = hoja[rango]
            celda.value = texto
            celda.border = borde_negro
            if rango.startswith("B"):
                celda.alignment = Alignment(horizontal="left")
            else:
                celda.alignment = Alignment(horizontal="center", vertical="center")
                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = color_fondo
        for rango, texto in datoscombinados:
            hoja.merge_cells(rango)
            celda = hoja[rango.split(":")[0]]
            celda.value = texto
            celda.alignment = Alignment(horizontal="center", vertical="center")
        # Agregar bordes a las celdas fusionadas
        for rango in ["A1:A4", "B1:E4", "A6:E7", "D8:E8", "D9:E9", "D10:E10", "D11:E11"]:
            for fila_celdas in hoja[rango]:
                for celda in fila_celdas:
                    celda.border = borde_negro
        # Fila de encabezados detallados
        encabezados = ["Fecha", "Hora", "Profundidad (m)", "Impedancia", "Observación"]
        for col, encabezado in enumerate(encabezados, 1):
            hoja[chr(64 + col) + "13"].value = encabezado
            celda = hoja[chr(64 + col) + "13"]
            celda.font = Font(bold=True, color="FFFFFF")
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.fill = color_fondo
            celda.border = borde_negro
    
    def exportarExcelCotasTerreno(idproyecto, nameproyecto, idzona, terrenosmarcados):
        respuesta = EmpresaConfiguracion.obtenerDataEmpresa()
        logo = respuesta[4]
        libro = Workbook()
        cont = 0
        for namecota, idinstrumento, idcota in terrenosmarcados:
            datospluvio = DatosController.ctrlListarDataCotaTerreno(idproyecto, idzona, idinstrumento)
            if datospluvio:
                if cont == 0:
                    hoja = libro.active
                    hoja.title = namecota
                else:
                    hoja = libro.create_sheet(title=namecota)
                # Configurar la hoja (común para todas)
                ExportarData.configurarCabeceraHojaCotaTerreno(hoja, nameproyecto, namecota, logo)
                # Insertar datos en la tabla comenzando desde la fila 12
                for fila in datospluvio:
                    hoja.append(list(fila))
                cont += 1
        rutaexcel = resource_path("resources/workspace/dataequipos.xlsx")
        libro.save(rutaexcel)
        archivo_destino, _ = QFileDialog.getSaveFileName(None, "Guardar Excel en", "Terrenos", "Archivos de Excel (*.xlsx);;Todos los archivos (*)")
        if archivo_destino:
            # Asegurarse de que el archivo tenga la extensión .xlsx
            if not archivo_destino.lower().endswith('.xlsx'):
                archivo_destino += '.xlsx'
            try:
                shutil.copy(rutaexcel, archivo_destino)
                mostrar_mensaje("Data Exportada", f"El Excel se ha guardado en: {archivo_destino}", "informacion")
            except Exception as e:  # Captura cualquier excepción
                mostrar_mensaje("Error al Exportar", f"No se pudo guardar el Excel: {str(e)}", "advertencia")
    
    def configurarCabeceraHojaCotaTerreno(hoja, proyectoname, namepluvio, logo):
        # Definir colores, bordes, y otros atributos comunes
        color_fondo = PatternFill(start_color="3C3C3C", end_color="3C3C3C", fill_type="solid")
        borde_negro = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        # Ajustar el ancho de las columnas para mejor presentación
        for col in range(1, 5):  # 4 columnas de ancho
            hoja.column_dimensions[chr(64 + col)].width = 20
        # Ajustar el alto de las filas de 1 a 5 para el área de la imagen
        for row in range(1, 5):
            hoja.row_dimensions[row].height = 25
        # Ruta de la imagen
        if logo:
            imagen_stream = MetodosGenerales.convertirBlobImagen(logo)
            imagen = ExcelImage(imagen_stream)
        else:
            ui_file_path = resource_path("resources/logo.png")
            imagen = ExcelImage(ui_file_path)
        imagen.width = 132
        imagen.height = 132
        imagen.anchor = "A1"
        hoja.merge_cells("A1:A4")
        hoja.add_image(imagen)
        # Agregar el título
        hoja.merge_cells("B1:D4")
        celda_titulo = hoja["B1"]
        celda_titulo.value = f"DATA DE COTA DE TERRENO - {proyectoname.upper()}"
        celda_titulo.font = Font(size=18, bold=True)
        celda_titulo.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Agregar bordes a las celdas fusionadas
        for rango in ["A1:D4", "A6:D6"]:
            for fila_celdas in hoja[rango]:
                for celda in fila_celdas:
                    celda.border = borde_negro
        # Fila de encabezados detallados
        encabezados = ["Fecha", "Hora", "Cota (msnm)", "Observación"]
        for col, encabezado in enumerate(encabezados, 1):
            hoja[chr(64 + col) + "6"].value = encabezado
            celda = hoja[chr(64 + col) + "6"]
            celda.font = Font(bold=True, color="FFFFFF")
            celda.alignment = Alignment(horizontal="center", vertical="center")
            celda.fill = color_fondo
            celda.border = borde_negro
    
    def descargarFormatoExcel(tipo):
        if tipo == "prisma":
            rutaexcel = resource_path("resources/formatos/FormatoPrismas.xlsx")
            nombrearchivo = "FormatoPrismas.xlsx"
        elif tipo == "cuerda":
            rutaexcel = resource_path("resources/formatos/FormatoPiezometroCuerda.xlsx")
            nombrearchivo = "FormatoPiezometrosCuerda.xlsx"
        elif tipo == "casagrande":
            rutaexcel = resource_path("resources/formatos/FormatoPiezometroCasagrande.xlsx")
            nombrearchivo = "FormatoPiezometrosCasagrande.xlsx"
        elif tipo == "celda":
            rutaexcel = resource_path("resources/formatos/FormatoCeldas.xlsx")
            nombrearchivo = "FormatoCeldas.xlsx"
        elif tipo == "pluvio":
            rutaexcel = resource_path("resources/formatos/FormatoPluviometros.xlsx")
            nombrearchivo = "FormatoPluviometros.xlsx"
        elif tipo == "cota":
            rutaexcel = resource_path("resources/formatos/FormatoCotaTerreno.xlsx")
            nombrearchivo = "FormatoCotasTerreno.xlsx"
        elif tipo == "sondaje":
            rutaexcel = resource_path("resources/formatos/FormatoTDR.xlsx")
            nombrearchivo = "FormatoTDR.xlsx"
        else:
            rutaexcel = resource_path("resources/formatos/FormatoAcelerografos.xlsx")
            nombrearchivo = "FormatoAcelerografos.xlsx"
        archivo_destino, _ = QFileDialog.getSaveFileName(None, "Guardar Excel en", nombrearchivo, "Archivos de Excel (*.xlsx);;Todos los archivos (*)")
        if archivo_destino:
            # Asegurarse de que el archivo tenga la extensión .xlsx
            if not archivo_destino.lower().endswith('.xlsx'):
                archivo_destino += '.xlsx'
            try:
                shutil.copy(rutaexcel, archivo_destino)
                mostrar_mensaje("Data Exportada", f"El Excel se ha guardado en: {archivo_destino}", "informacion")
            except Exception as e:  # Captura cualquier excepción
                mostrar_mensaje("Error al Exportar", f"No se pudo guardar el Excel: {str(e)}", "advertencia")
    