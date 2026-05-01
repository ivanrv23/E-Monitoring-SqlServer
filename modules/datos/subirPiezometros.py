import pandas as pd
from openpyxl import load_workbook
from PySide6.QtWidgets import (QDialog, QVBoxLayout, QComboBox, QFileDialog, QCheckBox, QPushButton, QTableWidget, QDoubleSpinBox,
                        QSpinBox, QTreeWidgetItem, QTreeWidget, QFormLayout, QDialogButtonBox, QMessageBox, QLabel, QTextEdit,
                        QLineEdit, QTableView)
from PySide6.QtUiTools import QUiLoader
from PySide6.QtCore import Qt
from datetime import datetime, time
from utils.common.rutasarchivos import resource_path
from utils.generic.cargariconos import cargarIcono
from utils.common.alertas import mostrar_mensaje
from utils.shared.pegarDatosTabla import configurar_tabla_para_pegado
from utils.generic.listaiconos import ListaIconos
from utils.shared.arbolmarcado import TreeCheckbox
from utils.common.metodosGenerales import MetodosGenerales
from controllers.PiezometroController import PiezometroController
from controllers.ProyectoController import ProyectoController
from controllers.InterfazController import InterfazController

class SubirPiezometros:
    
    def cargarPiezometrosCuerda(main, proyectoid):
        loaderLoading = QUiLoader()        
        ui_file_path = resource_path("ui/datapiezometrocuerda.ui")
        ui_file = loaderLoading.load(ui_file_path, None)
        dialogoPiezocuerda = QDialog()
        dialogoPiezocuerda.setWindowTitle("Data Piezómetros Cuerda Vibrante")
        layout_piezocuerda = QVBoxLayout()
        layout_piezocuerda.addWidget(ui_file)
        dialogoPiezocuerda.setLayout(layout_piezocuerda)
        # tools
        comboPiezometros = dialogoPiezocuerda.findChild(QComboBox, "combo_piezometros")
        tabladata = dialogoPiezocuerda.findChild(QTableWidget, "table_piezometros_calculo")
        checknivel = dialogoPiezocuerda.findChild(QCheckBox, "check_elevacion")
        lblrespuesta = dialogoPiezocuerda.findChild(QLabel, "label_mensaje_estado")
        # agregar una celda a cada tabla
        row_position1 = tabladata.rowCount()
        tabladata.insertRow(row_position1)
        botonGuardarData = dialogoPiezocuerda.findChild(QPushButton, "btn_guardar_piezometros")
        # cargar piezómetros en el combo
        listapiezometros = PiezometroController.ctrlListarPiezometrosCuerda(proyectoid)
        if listapiezometros:
            for fila in listapiezometros:
                comboPiezometros.addItem(str(fila[3]), fila[0])
        else:
            comboPiezometros.addItem("Sin Piezómetros")
            comboPiezometros.setEnabled(False)
            botonGuardarData.setEnabled(False)
        configurar_tabla_para_pegado(tabladata)
        # GUARDAR DATA CUERDA VIBRANTE TABLA
        def guardarPiezometrosCuerdaTabla():
            idpiezometro = comboPiezometros.currentData()
            estadonivel = checknivel.isChecked()
            filas = tabladata.rowCount()
            if filas > 0 and proyectoid != 0:
                data = []
                estado = False
                for row in range(filas):
                    datosfila = []
                    datosfila.append(idpiezometro)
                    fila_valida = True
                    c = 0
                    for column in range(tabladata.columnCount()):
                        item = tabladata.item(row, column)
                        valor = item.text().strip() if item else ""
                        if valor != "":
                            if column == 0:
                                valor = MetodosGenerales.validarFormatoFecha(valor)
                                if not valor:
                                    mensaje = "La fecha no tiene un formato adecuado."
                                    fila_valida = False
                                    break
                            elif column == 1:
                                valor = MetodosGenerales.validarFormatoHora(valor)
                                if not valor:
                                    mensaje = "La hora no tiene un formato adecuado."
                                    fila_valida = False
                                    break
                            elif column > 1 and column < 6:
                                if not MetodosGenerales.validarEsNumero(valor):
                                    mensaje = "Algunas lecturas no son numéricas."
                                    fila_valida = False
                                    break
                        else:
                            if column == 0:
                                fila_valida = False
                                mensaje = "La fecha está vacía."
                            elif column == 1:
                                valor = "00:00:00"
                            elif column == 2:
                                fila_valida = False
                                mensaje = "La frecuencia está vacía."
                            elif column == 3:
                                fila_valida = False
                                mensaje = "La temperatura está vacía."
                            elif column == 4:
                                valor = 0
                            elif column == 5:
                                valor = 0
                            c += 1
                        datosfila.append(valor)
                    if c != 7:
                        if fila_valida and len(datosfila) == 8:
                            data.append(datosfila)
                            estado = True
                        else:
                            estado = False
                            break
                if estado:
                    respuesta = PiezometroController.ctrlGuardarPiezometrosCuerdaCalculada(proyectoid, data, estadonivel)
                    if respuesta:
                        lblrespuesta.setText('Guardado correctamente')
                        lblrespuesta.setStyleSheet("color: green;")
                        # Limpiar filas tabla
                        tabladata.setRowCount(0)
                        tabladata.insertRow(0)
                        # actualizar árbol checkbox
                        data = PiezometroController.ctrlTraerDataPiezometro(idpiezometro, "PIEZOMETROCUERDA")
                        if data:
                            idinstrumento, idcomponente, nombrezona = data[0], data[1], data[2]
                            treewidgetdatos = main.findChild(QTreeWidget, "tree_actual_datos")
                            treewidgetvisor = main.findChild(QTreeWidget, "tree_actual_visor")
                            treewidgetpiezo = main.findChild(QTreeWidget, "tree_actual_piezometros")
                            TreeCheckbox.eliminarCheckbox(treewidgetdatos, "Piezómetros Cuerda Vibrante", idinstrumento, "piezometrocuerda")
                            TreeCheckbox.eliminarCheckbox(treewidgetvisor, "Piezómetros Cuerda Vibrante", idinstrumento, "piezometrocuerda")
                            TreeCheckbox.eliminarCheckbox(treewidgetpiezo, "Piezómetros Cuerda Vibrante", idinstrumento, "piezometrocuerda")
                            # Crear piezometro cuerda en nuevo componente
                            piezocu = InterfazController.ctrlListarComponentePiezometroCuerda(idinstrumento)
                            if piezocu:
                                TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidgetdatos, nombrezona, idcomponente, proyectoid, "Piezómetros Cuerda Vibrante", "3", piezocu, "piezometrocuerda")
                                TreeCheckbox.crearNuevoGrupoCheckboxesDoble(treewidgetpiezo, nombrezona, idcomponente, proyectoid, "Piezómetros Cuerda Vibrante", "1", piezocu, "piezometrocuerda")
                                fechas = InterfazController.ctrlListarFechasPiezometroCodigo("Automatizado", idcomponente, idinstrumento, proyectoid)
                                if fechas:
                                    ultima_fecha = fechas[-1][0]
                                    piezo = piezocu[0]
                                    piezometros = [(piezo[0], piezo[1], piezo[2], piezo[3], piezo[4], piezo[5], piezo[6], ultima_fecha)]
                                    TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidgetvisor, nombrezona, idcomponente, proyectoid, "Piezómetros Cuerda Vibrante", "4", piezometros, "piezometrocuerda", "SI")
                    else:
                        lblrespuesta.setText(f"En la fila {len(data) + 1}: {mensaje}")
                        lblrespuesta.setStyleSheet("color: red;")
                else:
                    lblrespuesta.setText(f"En la fila {len(data) + 1}: {mensaje}")
                    lblrespuesta.setStyleSheet("color: orange;")  
                    
        # Conectar señales
        botonGuardarData.clicked.connect(guardarPiezometrosCuerdaTabla)
        dialogoPiezocuerda.exec()
    
    def cargarDataFormatosCuerda(main, idproyecto, tipo):
        loader = QUiLoader()
        ui_file_path = resource_path("ui/cargardataformato.ui")
        ui_file = loader.load(ui_file_path, None)
        dialogo = QDialog()
        dialogo.setWindowTitle("Data Piezómetros Cuerda Vibrante")
        layout_procesar_data = QVBoxLayout()
        layout_procesar_data.addWidget(ui_file)
        dialogo.setLayout(layout_procesar_data)
        ruta = "resources/iconos/fontawesome/solid/file-arrow-up.svg"
        botonSubir = dialogo.findChild(QPushButton, "btn_cargar_archivo")
        cargarIcono(botonSubir, ruta)
        ubicacion_archivo = dialogo.findChild(QLineEdit, "input_archivo")
        ubicacion_archivo.setReadOnly(True)
        comboComponentes = dialogo.findChild(QComboBox, "combo_componentes")
        labelRespuesta = dialogo.findChild(QLabel, "label_mensaje")
        botonAceptar = dialogo.findChild(QPushButton, "btn_aceptar")
        componentes = ProyectoController.ctrlObtenerComponentesProyecto(idproyecto)
        if len(componentes) > 0:
            for fila in componentes:
                comboComponentes.addItem(str(fila[2]), fila[0])
            comboComponentes.setEnabled(True)
        else:
            comboComponentes.addItem("Sin Componentes")
            comboComponentes.setEnabled(False)
            botonAceptar.setEnabled(False)
        def cargar_archivo():
            file_names, _ = QFileDialog.getOpenFileNames(None, "Cargar Archivos", "", "Archivos Excel (*.xlsx)")
            if file_names:
                ubicacion_archivo.setText("\n".join(file_names))
        def procesar_archivo():
            if not ubicacion_archivo.text().strip():
                labelRespuesta.setText("No se cargó ningún archivo.")
                labelRespuesta.setStyleSheet("color: red;")
                return
            else:
                try:
                    idcompo = comboComponentes.currentData()
                    if tipo == "FORMATO":
                        respuesta, equipos, erroneos = SubirPiezometros.registrarDataPiezometrosCuerda(idproyecto, ubicacion_archivo.text(), idcompo)
                    else:
                        respuesta, equipos, erroneos = SubirPiezometros.registrarDataExcelPiezometrosCuerda(idproyecto, ubicacion_archivo.text(), idcompo)
                    if respuesta:
                        ubicacion_archivo.clear()
                        if len(erroneos) > 0:
                            labelRespuesta.setText(f"Archivos erróneos: {erroneos}")
                        else:
                            labelRespuesta.setText("Guardado correctamente.")
                        labelRespuesta.setStyleSheet("color: green;")
                        # actualizar árbol checkbox
                        for idpiezometro in equipos:
                            data = PiezometroController.ctrlTraerDataPiezometro(idpiezometro, "PIEZOMETROCUERDA")
                            if data:
                                idinstrumento, idcomponente, nombrezona = data[0], data[1], data[2]
                                treewidgetdatos = main.findChild(QTreeWidget, "tree_actual_datos")
                                treewidgetvisor = main.findChild(QTreeWidget, "tree_actual_visor")
                                treewidgetpiezo = main.findChild(QTreeWidget, "tree_actual_piezometros")
                                TreeCheckbox.eliminarCheckbox(treewidgetdatos, "Piezómetros Cuerda Vibrante", idinstrumento, "piezometrocuerda")
                                TreeCheckbox.eliminarCheckbox(treewidgetvisor, "Piezómetros Cuerda Vibrante", idinstrumento, "piezometrocuerda")
                                TreeCheckbox.eliminarCheckbox(treewidgetpiezo, "Piezómetros Cuerda Vibrante", idinstrumento, "piezometrocuerda")
                                # Crear piezometro cuerda en nuevo componente
                                piezocu = InterfazController.ctrlListarComponentePiezometroCuerda(idinstrumento)
                                if piezocu:
                                    TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidgetdatos, nombrezona, idcomponente, idproyecto, "Piezómetros Cuerda Vibrante", "3", piezocu, "piezometrocuerda")
                                    TreeCheckbox.crearNuevoGrupoCheckboxesDoble(treewidgetpiezo, nombrezona, idcomponente, idproyecto, "Piezómetros Cuerda Vibrante", "1", piezocu, "piezometrocuerda")
                                    fechas = InterfazController.ctrlListarFechasPiezometroCodigo("Automatizado", idcomponente, idinstrumento, idproyecto)
                                    if fechas:
                                        ultima_fecha = fechas[-1][0]
                                        piezo = piezocu[0]
                                        piezometros = [(piezo[0], piezo[1], piezo[2], piezo[3], piezo[4], piezo[5], piezo[6], ultima_fecha)]
                                        TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidgetvisor, nombrezona, idcomponente, idproyecto, "Piezómetros Cuerda Vibrante", "4", piezometros, "piezometrocuerda", "SI")
                    else:
                        if len(erroneos) > 0:
                            labelRespuesta.setText(f"Error en los archivos: {erroneos}")
                        else:
                            labelRespuesta.setText("No se guardó la data.")
                        labelRespuesta.setStyleSheet("color: red;")
                except ValueError as e:
                    labelRespuesta.setText(str(e))
                    labelRespuesta.setStyleSheet("color: red;")
        botonSubir.clicked.connect(cargar_archivo)
        botonAceptar.clicked.connect(procesar_archivo)
        dialogo.exec()
    
    def registrarDataPiezometrosCuerda(proyectoid, ubicacion, idcomponente):
        erroneos = []
        data = []
        equipos = []
        respuesta = False
        encabezadomb = ['Fecha', 'Hora', 'Frecuencia (Dg)', 'Temperatura (°C)', 'Presión (mb)', 'mca (m)', 'Observación']
        encabezadokpa = ['Fecha', 'Hora', 'Frecuencia (Dg)', 'Temperatura (°C)', 'Presión (kPa)', 'mca (m)', 'Observación']
        archivos = ubicacion.split("\n")
        for file_name in archivos:
            file_name = file_name.strip()
            if not file_name or not file_name.endswith('.xlsx'):
                continue
            try:
                df_header = pd.read_excel(file_name, header=None, nrows=1, skiprows=13, engine='openpyxl')
                encabezado_archivomb = [str(col).strip() for col in df_header.iloc[0, :len(encabezadomb)]]
                encabezado_archivokpa = [str(col).strip() for col in df_header.iloc[0, :len(encabezadokpa)]]
                if encabezado_archivomb == encabezadomb:
                    unidadpresion = "mb"
                elif encabezado_archivokpa == encabezadokpa:
                    unidadpresion = "kPa"
                else:
                    erroneos.append(file_name.split("/")[-1])
                    continue
                wb = load_workbook(file_name, data_only=True)
                hoja = wb.active
                # obtener data general
                nombrepiezo = hoja["B8"].value
                seriepiezo = hoja["B9"].value
                coordeste = hoja["B10"].value
                coordnorte = hoja["B11"].value
                instalacion = hoja["B12"].value
                fundacion = hoja["B13"].value
                superficie = hoja["E8"].value
                inclinacion = hoja["E9"].value
                azimuth = hoja["E10"].value
                cf = hoja["E11"].value
                tk = hoja["E12"].value
                frecuenciaini = hoja["E13"].value
                temperaini = hoja["G8"].value
                presionini = hoja["G9"].value
                conversion = hoja["G10"].value
                constantea = hoja["G11"].value
                constanteb = hoja["G12"].value
                constantec = hoja["G13"].value
                comentario = ""
                wb.close()
                # Validar datos
                if pd.isna(nombrepiezo) or proyectoid == 0 or not idcomponente:
                    erroneos.append(file_name.split("/")[-1])
                    continue
                idpiezometro = None
                respu, info = PiezometroController.ctrlComprobarExisteNombrePiezometro(proyectoid, nombrepiezo, "Automatizado")
                if respu:
                    idpiezometro = info[0]
                    if float(coordeste) != 0 and float(coordnorte) != 0:
                        datos = (seriepiezo, coordeste, coordnorte, instalacion, fundacion, inclinacion, azimuth, cf, tk, frecuenciaini, temperaini, presionini, constantea, constanteb, constantec, conversion, comentario, idpiezometro)
                        response = PiezometroController.ctrlActualizarPiezometroCuerdaFormato(datos)
                else:
                    if pd.isna(superficie):
                        continue
                    if pd.isna(coordnorte):
                        coordnorte = 0
                    else:
                        try:
                            coordnorte = float(coordnorte)
                        except ValueError:
                            coordnorte = 0
                    if pd.isna(coordeste):
                        coordeste = 0
                    else:
                        try:
                            coordeste = float(coordeste)
                        except ValueError:
                            coordeste = 0
                    if pd.isna(instalacion):
                        instalacion = 0
                    else:
                        try:
                            instalacion = float(instalacion)
                        except ValueError:
                            instalacion = 0
                    if pd.isna(fundacion):
                        fundacion = 0
                    else:
                        try:
                            fundacion = float(fundacion)
                        except ValueError:
                            fundacion = 0
                    if pd.isna(inclinacion):
                        inclinacion = 90
                    else:
                        try:
                            inclinacion = float(inclinacion)
                        except ValueError:
                            inclinacion = 90
                    if pd.isna(azimuth):
                        azimuth = 0
                    else:
                        try:
                            azimuth = float(azimuth)
                        except ValueError:
                            azimuth = 0
                    if pd.isna(conversion):
                        conversion = 0
                    else:
                        try:
                            conversion = float(conversion)
                        except ValueError:
                            conversion = 0
                    if pd.isna(cf):
                        cf = 0
                    else:
                        try:
                            cf = float(cf)
                        except ValueError:
                            cf = 0
                    if pd.isna(tk):
                        tk = 0
                    else:
                        try:
                            tk = float(tk)
                        except ValueError:
                            tk = 0
                    fecha = f"{datetime.now().strftime('%Y-%m-%d')} 00:00:00"
                    datos = (proyectoid, nombrepiezo, seriepiezo, coordeste, coordnorte, instalacion, fundacion, inclinacion, azimuth, cf, tk, frecuenciaini, temperaini, presionini, unidadpresion, constantea, constanteb, constantec, conversion, comentario)
                    respues = PiezometroController.ctrlRegistrarPiezometroCuerdaFormato(idcomponente, datos, fecha, superficie, "PCV")
                    if respues:
                        idpiezometro = respues
                if idpiezometro is not None:
                    df = pd.read_excel(file_name, header=None, skiprows=14, engine='openpyxl')
                    df.columns = ['fecha', 'hora', 'frecuencia', 'temperatura', 'presion', 'mca', 'observacion']
                    for _, row in df.iterrows():
                        fecha = row['fecha']
                        hora = row['hora']
                        mca = row['mca']
                        observacion = row['observacion']
                        if pd.isna(fecha) or pd.isna(mca):
                            continue
                        # Manejo de la columna 'fecha'
                        if isinstance(fecha, (pd.Timestamp, datetime)):
                            fecha = fecha.date()  # Convertir a date
                            fecha = fecha.strftime('%Y-%m-%d')
                        elif isinstance(fecha, str):
                            fecha = MetodosGenerales.validarFormatoFecha(fecha)
                            if fecha is None:
                                continue
                        else:
                            continue
                        # Manejo de la columna 'hora'
                        if isinstance(hora, (pd.Timestamp, datetime)):
                            hora = hora.time()
                            hora = hora.strftime('%H:%M:%S')
                        elif isinstance(hora, time):
                            hora = hora.strftime('%H:%M:%S')
                        elif isinstance(hora, str):
                            hora = MetodosGenerales.validarFormatoHora(hora) or "00:00:00"
                        else:
                            hora = "00:00:00"
                        try:
                            mca = float(mca)
                        except (ValueError, TypeError):
                            continue
                        frecu = float(row['frecuencia']) if not pd.isna(row['frecuencia']) else 0
                        tempe = float(row['temperatura']) if not pd.isna(row['temperatura']) else 0
                        presio = float(row['presion']) if not pd.isna(row['presion']) else 0
                        observa = "" if pd.isna(observacion) else str(observacion).strip()
                        data.append((idpiezometro, fecha, hora, frecu, tempe, presio, mca, observa))
                    if data:
                        respon = PiezometroController.ctrlGuardarPiezometrosCuerdaCalculada(proyectoid, data, False)
                        if respon:
                            equipos.append(idpiezometro)
                            respuesta = True
                        else:
                            erroneos.append(file_name.split("/")[-1])
                    else:
                        erroneos.append(file_name.split("/")[-1])
                else:
                    erroneos.append(file_name.split("/")[-1])
            except Exception:
                erroneos.append(file_name.split("/")[-1])
        return respuesta, equipos, erroneos
    
    def registrarDataExcelPiezometrosCuerda(proyectoid, ubicacion, idcomponente):
        erroneos = []
        data = []
        equipos = []
        respuesta = False
        # Definir encabezado esperado y las celdas donde debe estar cada columna
        encabezado_esperado = {
            'A12': 'Fecha',
            'D12': 'Frecuencia (Digits)',
            'I12': 'Frecuencia (Hz)',
            'M12': 'Temperatura (°C)',
            'R12': 'Presión (MPa)',
            'V12': '(mca)',
            'Y12': 'Cota piezométrica    (m s.n.m.)',
            'AD12': 'Observación'
        }
        archivos = ubicacion.split("\n")
        for file_name in archivos:
            file_name = file_name.strip()
            if not file_name or not file_name.endswith('.xlsx'):
                continue
            try:
                # Obtener todas las hojas del archivo Excel
                wb = load_workbook(file_name, data_only=True)
                sheet_names = wb.sheetnames
                # Procesar cada hoja
                for sheet_name in sheet_names:
                    try:
                        # Seleccionar la hoja actual
                        hoja = wb[sheet_name]
                        # Validar el encabezado leyendo celdas específicas
                        encabezado_valido = True
                        for celda, valor_esperado in encabezado_esperado.items():
                            valor_celda = hoja[celda].value
                            if valor_celda is None:
                                valor_celda = ""
                            valor_celda = str(valor_celda).strip()
                            if valor_celda != valor_esperado:
                                encabezado_valido = False
                                break
                        if not encabezado_valido:
                            continue
                        # Obtener data general de la hoja actual
                        nombrepiezo = hoja["C6"].value
                        codigopiezo = hoja["C8"].value
                        seriepiezo = hoja["C9"].value
                        comentario = hoja["C10"].value
                        instalacion = hoja["P7"].value
                        fundacion = hoja["P7"].value
                        coordeste = hoja["N9"].value
                        coordnorte = hoja["N10"].value
                        superficie = hoja["P6"].value
                        cf = hoja["W10"].value
                        tk = hoja["AE10"].value
                        inclinacion = 90
                        azimuth = 0
                        frecuini = hoja["Y5"].value
                        temperaini = hoja["AD5"].value
                        presionini = 0
                        constantea = hoja["W8"].value
                        constanteb = hoja["AA8"].value
                        constantec = hoja["AE8"].value
                        conversion = 101.97
                        # Validar datos
                        if pd.isna(nombrepiezo) or proyectoid == 0 or not idcomponente:
                            continue  # Continuar con la siguiente hoja
                        idpiezometro = None
                        respu, info = PiezometroController.ctrlComprobarExisteNombrePiezometro(proyectoid, nombrepiezo, "Automatizado")
                        if respu:
                            idpiezometro = info[0]
                        else:
                            if pd.isna(superficie):
                                continue
                            # Procesamiento de datos con validaciones
                            if pd.isna(coordnorte):
                                coordnorte = 0
                            else:
                                try:
                                    coordnorte = float(coordnorte)
                                except ValueError:
                                    coordnorte = 0
                            if pd.isna(coordeste):
                                coordeste = 0
                            else:
                                try:
                                    coordeste = float(coordeste)
                                except ValueError:
                                    coordeste = 0
                            if pd.isna(instalacion):
                                instalacion = 0
                            else:
                                try:
                                    instalacion = float(instalacion)
                                except ValueError:
                                    instalacion = 0
                            if pd.isna(fundacion):
                                fundacion = 0
                            else:
                                try:
                                    fundacion = float(fundacion)
                                except ValueError:
                                    fundacion = 0
                            if pd.isna(cf):
                                cf = 0
                            else:
                                try:
                                    cf = float(cf)
                                except ValueError:
                                    cf = 0
                            if pd.isna(tk):
                                tk = 0
                            else:
                                try:
                                    tk = float(tk)
                                except ValueError:
                                    tk = 0
                            fecha = f"{datetime.now().strftime('%Y-%m-%d')} 00:00:00"
                            datos = (proyectoid, nombrepiezo, seriepiezo, coordeste, coordnorte, instalacion, fundacion, inclinacion, azimuth, cf, tk, frecuini, temperaini, presionini, "Dg", constantea, constanteb, constantec, conversion, comentario)
                            respues = PiezometroController.ctrlRegistrarPiezometroCuerdaFormato(idcomponente, datos, fecha, superficie, "PCV")
                            if respues:
                                idpiezometro = respues
                                equipos.append(idpiezometro)
                        if idpiezometro is not None:
                            # Leer datos desde la fila 13 en adelante usando las columnas específicas
                            sheet_data = []
                            fila = 13  # Empezar desde la fila 13 (debajo del encabezado)
                            while True:
                                # Leer datos de la fila actual
                                fecha_valor = hoja[f'A{fila}'].value
                                digits_valor = hoja[f'D{fila}'].value
                                frecuencia_valor = hoja[f'I{fila}'].value
                                temperatura_valor = hoja[f'M{fila}'].value
                                presion_valor = hoja[f'R{fila}'].value
                                mca_valor = hoja[f'V{fila}'].value
                                cota_valor = hoja[f'Y{fila}'].value
                                observacion_valor = hoja[f'AD{fila}'].value
                                # Si no hay fecha o mca, terminar la lectura
                                if pd.isna(fecha_valor) or pd.isna(mca_valor):
                                    # Verificar si es una fila completamente vacía
                                    if all(pd.isna(val) for val in [fecha_valor, digits_valor, frecuencia_valor, temperatura_valor, presion_valor, mca_valor, cota_valor, observacion_valor]):
                                        break
                                    # Si solo faltan fecha o mca, saltar esta fila
                                    fila += 1
                                    continue
                                # Manejo de la columna 'fecha'
                                if isinstance(fecha_valor, (pd.Timestamp, datetime)):
                                    fecha_procesada = fecha_valor.date()
                                    fecha_procesada = fecha_procesada.strftime('%Y-%m-%d')
                                elif isinstance(fecha_valor, str):
                                    fecha_procesada = MetodosGenerales.validarFormatoFecha(fecha_valor)
                                    if fecha_procesada is None:
                                        fila += 1
                                        continue
                                else:
                                    fila += 1
                                    continue
                                # Manejo de la columna 'hora'
                                hora = "00:00:00"
                                try:
                                    mca_procesada = float(mca_valor)
                                except (ValueError, TypeError):
                                    fila += 1
                                    continue
                                # Procesar otros valores numéricos
                                frecu = float(frecuencia_valor) if not pd.isna(frecuencia_valor) else 0
                                tempe = float(temperatura_valor) if not pd.isna(temperatura_valor) else 0
                                presio = float(presion_valor) if not pd.isna(presion_valor) else 0
                                # Procesar observación
                                observa = "" if pd.isna(observacion_valor) else str(observacion_valor).strip()
                                sheet_data.append((idpiezometro, fecha_procesada, hora, frecu, tempe, presio, mca_procesada, observa))
                                fila += 1
                            # Agregar datos de esta hoja al conjunto total
                            if sheet_data:
                                data.extend(sheet_data)
                    except Exception as e:
                        continue
                # Cerrar el workbook después de procesar todas las hojas
                wb.close()
                # Procesar todos los datos del archivo si hay datos válidos
                if data:
                    respon = PiezometroController.ctrlGuardarPiezometrosCuerdaCalculada(proyectoid, data, False)
                    if respon:
                        respuesta = True
                    else:
                        erroneos.append(file_name.split("/")[-1])
                else:
                    erroneos.append(file_name.split("/")[-1])
            except Exception as e:
                erroneos.append(file_name.split("/")[-1])
        return respuesta, equipos, erroneos
    
    def cargarPiezometrosCasagrande(main, proyectoid):
        loaderLoading = QUiLoader()        
        ui_file_path = resource_path("ui/datapiezometromanual.ui")
        ui_file = loaderLoading.load(ui_file_path, None)
        dialogoPiezomanual = QDialog()
        dialogoPiezomanual.setWindowTitle("Data Piezómetros Manuales")
        layout_piezomanual = QVBoxLayout()
        layout_piezomanual.addWidget(ui_file)
        dialogoPiezomanual.setLayout(layout_piezomanual)
        # tools
        comboPiezometros = dialogoPiezomanual.findChild(QComboBox, "combo_piezometros")
        tabladata = dialogoPiezomanual.findChild(QTableWidget, "table_piezometros_data")
        botonGuardarData = dialogoPiezomanual.findChild(QPushButton, "btn_guardar_piezometros")
        lblrespuesta = dialogoPiezomanual.findChild(QLabel, "label_mensaje_estado")
        # agregar una celda
        tabladata.insertRow(tabladata.rowCount())
        # cargar piezómetros en el combo
        listapiezometros = PiezometroController.ctrlListarPiezometrosManuales(proyectoid)
        if listapiezometros is not None:
            for fila in listapiezometros:
                comboPiezometros.addItem(str(fila[2]), fila[0])
        else:
            comboPiezometros.addItem("Sin Piezómetros")
            comboPiezometros.setEnabled(False)
        configurar_tabla_para_pegado(tabladata)
        # GUARDAR DATA
        def guardarPiezometrosManualesTabla():
            idpiezometro = comboPiezometros.currentData()
            filas = tabladata.rowCount()
            if filas > 0 and proyectoid != 0:
                data = []
                estado = False
                for row in range(filas):
                    datosfila = []
                    datosfila.append(idpiezometro)
                    fila_valida = False
                    c = 0
                    for column in range(tabladata.columnCount()):
                        item = tabladata.item(row, column)
                        valor = item.text().strip() if item else ""
                        mensaje = "Registrado correctamente."
                        if valor != "":
                            fila_valida = True
                            if column == 0:  # Validación de la fecha
                                valor = MetodosGenerales.validarFormatoFecha(valor)
                                if not valor:
                                    mensaje = "La fecha no tiene un formato adecuado."
                                    fila_valida = False
                                    break
                            elif column == 1:  # Validación de la hora
                                valor = MetodosGenerales.validarFormatoHora(valor)
                                if not valor:
                                    mensaje = "La hora no tiene un formato adecuado."
                                    fila_valida = False
                                    break
                            elif column == 2:  # Validación numérica de medidas
                                if not MetodosGenerales.validarEsNumero(valor):
                                    mensaje = "Las medidas deben ser numéricas."
                                    fila_valida = False
                                    break
                            datosfila.append(valor)
                        else:
                            c = c + 1
                            if column == 1:
                                valor = "00:00:00"
                                datosfila.append(valor)
                            elif column == 3:
                                valor = ""
                                datosfila.append(valor)
                            else:
                                fila_valida = False
                                mensaje = "Algunos campos están vacíos."
                    # Guardar la fila solo si es válida y tiene 4 columnas
                    if fila_valida and len(datosfila) == 5:
                        data.append(datosfila)
                        estado = True
                    else:
                        # Si todas son vacias omitir
                        if c == 4:
                            estado = True
                        else:
                            estado = False
                            break                    
                if estado:
                    respuesta = PiezometroController.ctrlGuardarPiezometrosManualesTabla(proyectoid, data)
                    if respuesta:
                        lblrespuesta.setText(mensaje)
                        lblrespuesta.setStyleSheet("color: green;")
                        # Limpiar filas tabla
                        tabladata.setRowCount(0)
                        tabladata.insertRow(0)
                        # actualizar árbol checkbox
                        data = PiezometroController.ctrlTraerDataPiezometro(idpiezometro, "PIEZOMETROMANUAL")
                        if data:
                            idinstrumento, idcomponente, nombrezona = data[0], data[1], data[2]
                            treewidgetdatos = main.findChild(QTreeWidget, "tree_actual_datos")
                            treewidgetvisor = main.findChild(QTreeWidget, "tree_actual_visor")
                            treewidgetpiezo = main.findChild(QTreeWidget, "tree_actual_piezometros")
                            TreeCheckbox.eliminarCheckbox(treewidgetdatos, "Piezómetros Casagrande", idinstrumento, "piezometromanual")
                            TreeCheckbox.eliminarCheckbox(treewidgetvisor, "Piezómetros Casagrande", idinstrumento, "piezometromanual")
                            TreeCheckbox.eliminarCheckbox(treewidgetpiezo, "Piezómetros Casagrande", idinstrumento, "piezometromanual")
                            # Crear piezometro manual en nuevo componente
                            piezocu = InterfazController.ctrlListarComponentePiezometroManual(idinstrumento)
                            if piezocu:
                                TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidgetdatos, nombrezona, idcomponente, proyectoid, "Piezómetros Casagrande", "4", piezocu, "piezometromanual")
                                TreeCheckbox.crearNuevoGrupoCheckboxesDoble(treewidgetpiezo, nombrezona, idcomponente, proyectoid, "Piezómetros Casagrande", "2", piezocu, "piezometromanual")
                                fechas = InterfazController.ctrlListarFechasPiezometroCodigo("Manual", idcomponente, idinstrumento, proyectoid)
                                if fechas:
                                    ultima_fecha = fechas[-1][0]
                                    piezo = piezocu[0]
                                    piezometros = [(piezo[0], piezo[1], piezo[2], piezo[3], piezo[4], piezo[5], piezo[6], ultima_fecha)]
                                    TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidgetvisor, nombrezona, idcomponente, proyectoid, "Piezómetros Casagrande", "5", piezometros, "piezometromanual", "SI")
                    else:
                        lblrespuesta.setText(f"Enla fila {len(data) + 1}: {mensaje}")
                        lblrespuesta.setStyleSheet("color: red;")
                else:
                    lblrespuesta.setText(f"Enla fila {len(data) + 1}: {mensaje}")
                    lblrespuesta.setStyleSheet("color: orange;")  
        # Conectar señal 
        botonGuardarData.clicked.connect(guardarPiezometrosManualesTabla)
        dialogoPiezomanual.exec()
    
    def cargarDataFormatosCasagrande(main, idproyecto):
        loader = QUiLoader()
        ui_file_path = resource_path("ui/cargardataformato.ui")
        ui_file = loader.load(ui_file_path, None)
        dialogo = QDialog()
        dialogo.setWindowTitle("Data Piezómetros casagrande")
        layout_procesar_data = QVBoxLayout()
        layout_procesar_data.addWidget(ui_file)
        dialogo.setLayout(layout_procesar_data)
        ruta = "resources/iconos/fontawesome/solid/file-arrow-up.svg"
        botonSubir = dialogo.findChild(QPushButton, "btn_cargar_archivo")
        cargarIcono(botonSubir, ruta)
        ubicacion_archivo = dialogo.findChild(QLineEdit, "input_archivo")
        ubicacion_archivo.setReadOnly(True)
        comboComponentes = dialogo.findChild(QComboBox, "combo_componentes")
        labelRespuesta = dialogo.findChild(QLabel, "label_mensaje")
        botonAceptar = dialogo.findChild(QPushButton, "btn_aceptar")
        componentes = ProyectoController.ctrlObtenerComponentesProyecto(idproyecto)
        if len(componentes) > 0:
            for fila in componentes:
                comboComponentes.addItem(str(fila[2]), fila[0])
            comboComponentes.setEnabled(True)
        else:
            comboComponentes.addItem("Sin Componentes")
            comboComponentes.setEnabled(False)
            botonAceptar.setEnabled(False)
        def cargar_archivo():
            file_names, _ = QFileDialog.getOpenFileNames(None, "Cargar Archivos", "", "Archivos Excel (*.xlsx)")
            if file_names:
                ubicacion_archivo.setText("\n".join(file_names))
        def procesar_archivo():
            if not ubicacion_archivo.text().strip():
                labelRespuesta.setText("No se cargó ningún archivo.")
                labelRespuesta.setStyleSheet("color: red;")
                return
            else:
                try:
                    idcompo = comboComponentes.currentData()
                    respuesta, equipos, erroneos = SubirPiezometros.registrarDataPiezometrosManuales(idproyecto, ubicacion_archivo.text(), idcompo)
                    if respuesta:
                        ubicacion_archivo.clear()
                        if len(erroneos) > 0:
                            labelRespuesta.setText(f"Archivos erróneos: {erroneos}")
                        else:
                            labelRespuesta.setText("Guardado correctamente.")
                        labelRespuesta.setStyleSheet("color: green;")
                        # actualizar árbol checkbox
                        for idpiezometro in equipos:
                            data = PiezometroController.ctrlTraerDataPiezometro(idpiezometro, "PIEZOMETROMANUAL")
                            if data:
                                idinstrumento, idcomponente, nombrezona = data[0], data[1], data[2]
                                treewidgetdatos = main.findChild(QTreeWidget, "tree_actual_datos")
                                treewidgetvisor = main.findChild(QTreeWidget, "tree_actual_visor")
                                treewidgetpiezo = main.findChild(QTreeWidget, "tree_actual_piezometros")
                                TreeCheckbox.eliminarCheckbox(treewidgetdatos, "Piezómetros Casagrande", idinstrumento, "piezometromanual")
                                TreeCheckbox.eliminarCheckbox(treewidgetvisor, "Piezómetros Casagrande", idinstrumento, "piezometromanual")
                                TreeCheckbox.eliminarCheckbox(treewidgetpiezo, "Piezómetros Casagrande", idinstrumento, "piezometromanual")
                                # Crear piezometro manual en nuevo componente
                                piezocu = InterfazController.ctrlListarComponentePiezometroManual(idinstrumento)
                                if piezocu:
                                    TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidgetdatos, nombrezona, idcomponente, idproyecto, "Piezómetros Casagrande", "4", piezocu, "piezometromanual")
                                    TreeCheckbox.crearNuevoGrupoCheckboxesDoble(treewidgetpiezo, nombrezona, idcomponente, idproyecto, "Piezómetros Casagrande", "2", piezocu, "piezometromanual")
                                    fechas = InterfazController.ctrlListarFechasPiezometroCodigo("Manual", idcomponente, idinstrumento, idproyecto)
                                    if fechas:
                                        ultima_fecha = fechas[-1][0]
                                        piezo = piezocu[0]
                                        piezometros = [(piezo[0], piezo[1], piezo[2], piezo[3], piezo[4], piezo[5], piezo[6], ultima_fecha)]
                                        TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidgetvisor, nombrezona, idcomponente, idproyecto, "Piezómetros Casagrande", "5", piezometros, "piezometromanual", "SI")
                    else:
                        if len(erroneos) > 0:
                            labelRespuesta.setText(f"Error en los archivos: {erroneos}")
                        else:
                            labelRespuesta.setText("No se guardó la data.")
                        labelRespuesta.setStyleSheet("color: red;")
                except ValueError as e:
                    labelRespuesta.setText(str(e))
                    labelRespuesta.setStyleSheet("color: red;")
        botonSubir.clicked.connect(cargar_archivo)
        botonAceptar.clicked.connect(procesar_archivo)
        dialogo.exec()
    
    def registrarDataPiezometrosManuales(proyectoid, ubicacion, idcomponente):
        erroneos = []
        data = []
        equipos = []
        respuesta = False
        encabezado = ['Fecha', 'Hora', 'Nivel Piezométrico (m)', 'Observación']
        archivos = ubicacion.split("\n")
        for file_name in archivos:
            file_name = file_name.strip()
            if not file_name or not file_name.endswith('.xlsx'):
                continue
            try:
                df_header = pd.read_excel(file_name, header=None, nrows=1, skiprows=20, engine='openpyxl')
                encabezados_archivo = [str(col).strip() for col in df_header.iloc[0, :len(encabezado)]]
                if encabezados_archivo != encabezado:
                    erroneos.append(file_name.split("/")[-1])
                    continue
                wb = load_workbook(file_name, data_only=True)
                hoja = wb.active
                nombrepiezo = hoja["C10"].value
                codigopiezo = hoja["C11"].value
                cotafondo = hoja["C12"].value
                fundacion = hoja["C13"].value
                coordeste = hoja["C14"].value
                coordnorte = hoja["C15"].value
                superficie = hoja["C16"].value
                inclinacion = hoja["C17"].value
                azimuth = hoja["C18"].value
                stickup = hoja["C19"].value
                comentario = hoja["C20"].value
                wb.close()
                if pd.isna(nombrepiezo) or proyectoid == 0 or not idcomponente:
                    erroneos.append(file_name.split("/")[-1])
                    continue
                idpiezometro = None
                respu, info = PiezometroController.ctrlComprobarExisteNombrePiezometro(proyectoid, nombrepiezo, "Manual")
                if respu:
                    idpiezometro = info[0]
                    if float(coordeste) != 0 and float(coordnorte) != 0:
                        datos = (codigopiezo, coordnorte, coordeste, cotafondo, fundacion, inclinacion, azimuth, stickup, comentario, idpiezometro)
                        response = PiezometroController.ctrlActualizarPiezometroManualFormato(datos)
                else:
                    if pd.isna(superficie):
                        continue
                    if pd.isna(coordnorte):
                        coordnorte = 0
                    else:
                        try:
                            coordnorte = float(coordnorte)
                        except ValueError:
                            coordnorte = 0
                    if pd.isna(coordeste):
                        coordeste = 0
                    else:
                        try:
                            coordeste = float(coordeste)
                        except ValueError:
                            coordeste = 0
                    if pd.isna(cotafondo):
                        cotafondo = 0
                    else:
                        try:
                            cotafondo = float(cotafondo)
                        except ValueError:
                            cotafondo = 0
                    if pd.isna(fundacion):
                        fundacion = 0
                    else:
                        try:
                            fundacion = float(fundacion)
                        except ValueError:
                            fundacion = 0
                    if pd.isna(inclinacion):
                        inclinacion = 90
                    else:
                        try:
                            inclinacion = float(inclinacion)
                        except ValueError:
                            inclinacion = 90
                    if pd.isna(azimuth):
                        azimuth = 0
                    else:
                        try:
                            azimuth = float(azimuth)
                        except ValueError:
                            azimuth = 0
                    if pd.isna(stickup):
                        stickup = 0
                    else:
                        try:
                            stickup = float(stickup)
                        except ValueError:
                            stickup = 0
                    fecha = f"{datetime.now().strftime('%Y-%m-%d')} 00:00:00"
                    datos = (proyectoid, nombrepiezo, codigopiezo, coordnorte, coordeste, cotafondo, fundacion, stickup, inclinacion, azimuth, comentario)
                    respues = PiezometroController.ctrlRegistrarPiezometroManualFormato(idcomponente, datos, fecha, superficie, "PVC")
                    if respues:
                        idpiezometro = respues
                if idpiezometro is not None:
                    df = pd.read_excel(file_name, header=None, skiprows=21, engine='openpyxl')
                    df.columns = ['fecha', 'hora', 'nivel', 'observacion']
                    for _, row in df.iterrows():
                        fecha = row['fecha']
                        hora = row['hora']
                        nivel = row['nivel']
                        observacion = row['observacion']
                        if pd.isna(fecha) or pd.isna(nivel):
                            continue
                        # Manejo de la columna 'fecha'
                        if isinstance(fecha, (pd.Timestamp, datetime)):
                            fecha = fecha.date()  # Convertir a date
                            fecha = fecha.strftime('%Y-%m-%d')
                        elif isinstance(fecha, str):
                            fecha = MetodosGenerales.validarFormatoFecha(fecha)
                            if fecha is None:
                                continue
                        else:
                            continue
                        # Manejo de la columna 'hora'
                        if isinstance(hora, (pd.Timestamp, datetime)):
                            hora = hora.time()
                            hora = hora.strftime('%H:%M:%S')
                        elif isinstance(hora, time):
                            hora = hora.strftime('%H:%M:%S')
                        elif isinstance(hora, str):
                            hora = MetodosGenerales.validarFormatoHora(hora) or "00:00:00"
                        else:
                            hora = "00:00:00"
                        try:
                            nivel = float(nivel)
                        except (ValueError, TypeError):
                            continue
                        observa = "" if pd.isna(observacion) else str(observacion).strip()
                        data.append((idpiezometro, fecha, hora, nivel, observa))
                    if data:
                        respon = PiezometroController.ctrlGuardarPiezometrosManualesTabla(proyectoid, data)
                        if respon:
                            equipos.append(idpiezometro)
                            respuesta = True
                        else:
                            erroneos.append(file_name.split("/")[-1])
                    else:
                        erroneos.append(file_name.split("/")[-1])
                else:
                    erroneos.append(file_name.split("/")[-1])
            except Exception:
                erroneos.append(file_name.split("/")[-1])
        return respuesta, equipos, erroneos
    
    def dialogoNuevaCotaPiezometrica(proyectoid):
        loaderLoading = QUiLoader()        
        ui_file_path = resource_path("ui/datacotapiezometrica.ui")
        ui_file = loaderLoading.load(ui_file_path, None)
        dialogo = QDialog()
        dialogo.setWindowTitle("Data Cotas Piezométricas")
        layout_piezo = QVBoxLayout()
        layout_piezo.addWidget(ui_file)
        dialogo.setLayout(layout_piezo)
        # tools
        comboPiezometros = dialogo.findChild(QComboBox, "combo_piezometros")
        tabladata = dialogo.findChild(QTableWidget, "table_cotas")
        botonGuardar = dialogo.findChild(QPushButton, "btn_guardar")
        botonCancelar = dialogo.findChild(QPushButton, "btn_cancelar")
        lblrespuesta = dialogo.findChild(QLabel, "label_mensaje")
        # agregar una celda
        tabladata.insertRow(tabladata.rowCount())
        # cargar piezómetros en el combo
        listapiezocuerda = PiezometroController.ctrlListarPiezometrosCuerda(proyectoid)
        listapiezomanual = PiezometroController.ctrlListarPiezometrosManuales(proyectoid)
        if listapiezocuerda is not None or listapiezomanual is not None:
            if listapiezocuerda is not None:
                for fila in listapiezocuerda:
                    comboPiezometros.addItem(str(fila[3]), (fila[0], "PCV"))
            if listapiezomanual is not None:
                for fila in listapiezomanual:
                    comboPiezometros.addItem(str(fila[2]), (fila[0], "PVC"))
        else:
            comboPiezometros.addItem("Sin Piezómetros")
            comboPiezometros.setEnabled(False)
            botonGuardar.setEnabled(False)
        configurar_tabla_para_pegado(tabladata)
        # GUARDAR DATA
        def guardarCotasTabla():
            infopiezometro = comboPiezometros.currentData()
            idpiezometro = infopiezometro[0]
            tipopiezometro = infopiezometro[1]
            filas = tabladata.rowCount()
            if filas > 0 and proyectoid is not None:
                data = []
                estado = False
                for row in range(filas):
                    datosfila = []
                    datosfila.append(idpiezometro)
                    datosfila.append(tipopiezometro)
                    fila_valida = False
                    c = 0
                    for column in range(tabladata.columnCount()):
                        item = tabladata.item(row, column)
                        valor = item.text().strip() if item else ""
                        mensaje = "Registrado correctamente."
                        if valor != "":
                            fila_valida = True
                            if column == 0:  # Validación de la fecha
                                valor = MetodosGenerales.validarFormatoFecha(valor)
                                if not valor:
                                    mensaje = "La fecha no tiene un formato adecuado."
                                    fila_valida = False
                                    break
                            elif column == 1:  # Validación numérica de medidas
                                if not MetodosGenerales.validarEsNumero(valor):
                                    mensaje = "Las cotas deben ser numéricas."
                                    fila_valida = False
                                    break
                            datosfila.append(valor)
                        else:
                            c = c + 1
                    # Guardar la fila solo si es válida y tiene 4 columnas
                    if fila_valida and len(datosfila) == 4:
                        data.append(datosfila)
                        estado = True
                    else:
                        # Si todas son vacias omitir
                        if c == 2:
                            estado = True
                        else:
                            estado = False
                            break                    
                if estado:
                    respuesta = PiezometroController.ctrlGuardarCotasPiezometricasTabla(data)
                    if respuesta:
                        lblrespuesta.setText(mensaje)
                        lblrespuesta.setStyleSheet("color: green;")
                        # Limpiar filas tabla
                        tabladata.setRowCount(0)
                        tabladata.insertRow(0)
                    else:
                        lblrespuesta.setText(f"Enla fila {len(data) + 1}: {mensaje}")
                        lblrespuesta.setStyleSheet("color: red;")
                else:
                    lblrespuesta.setText(f"Enla fila {len(data) + 1}: {mensaje}")
                    lblrespuesta.setStyleSheet("color: orange;")
        def cancelarCotas():
            dialogo.close()
        # Conectar señal 
        botonGuardar.clicked.connect(guardarCotasTabla)
        botonCancelar.clicked.connect(cancelarCotas)
        dialogo.exec()
    
    def cambiar_componente_piezocuerdas(idcomponente, idproyecto, treewidget, nombregrupo, tipogrupo, subgrupo, reiniciarvistas, vista="DATOS"):
        dialog = QDialog()
        dialog.setWindowTitle("Componente P. Cuerda Vibrante")
        layout = QFormLayout(dialog)
        # Campo componente
        label_titulo = QLabel("Componente:")
        combo_componente = QComboBox()
        label_mensaje = QLabel("")
        label_mensaje.setAlignment(Qt.AlignCenter)
        label_mensaje.setStyleSheet("QLabel { color: red; }")
        # cargar data componentes
        componentes = ProyectoController.ctrlObtenerComponentesProyecto(idproyecto)
        if componentes:
            for fila in componentes:
                combo_componente.addItem(str(fila[2]), fila[0])
            combo_componente.setCurrentIndex(combo_componente.findData(idcomponente))
        # Botones
        layout.addRow(label_titulo)
        layout.addRow(combo_componente)
        layout.addRow(label_mensaje)
        button_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        # Cambiar los textos a español
        button_box.button(QDialogButtonBox.Save).setText("Guardar")
        button_box.button(QDialogButtonBox.Cancel).setText("Cancelar")
        layout.addWidget(button_box)
        # Conectar los botones a las funciones correspondientes
        def actualizarDatos():
            componente = combo_componente.currentData()
            nombrezona = combo_componente.currentText()
            if str(idcomponente) == str(componente):
                dialog.reject()
            else:
                respuesta = PiezometroController.ctrlCambiarComponentePiezometrosCuerda(idcomponente, componente)
                if respuesta:
                    dialog.reject()
                    # Eliminar cuerdas
                    TreeCheckbox.eliminarCheckboxGrupo(treewidget, idcomponente, nombregrupo, tipogrupo)
                    # Crear cuerdas en nuevo componente
                    if vista == "PIEZOMETROS":
                        TreeCheckbox.crearNuevoGrupoCheckboxesDoble(treewidget, nombrezona, componente, idproyecto, nombregrupo, tipogrupo, respuesta, subgrupo)
                    else:
                        TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidget, nombrezona, componente, idproyecto, nombregrupo, tipogrupo, respuesta, subgrupo)
                    reiniciarvistas("Piezómetro")
                else:
                    label_mensaje.setText("Error al cambiar de componente.")
        button_box.accepted.connect(actualizarDatos)
        button_box.rejected.connect(dialog.reject)
        # Mostrar el diálogo
        dialog.setLayout(layout)
        dialog.exec()
    
    def eliminar_piezocuerdas(idproyecto, idzona, grupo, tipo, treewidget, reiniciarvistas):
        dlg = QMessageBox()
        dlg.setWindowTitle("Eliminar P. Cuerda Vibrante")
        dlg.setText(f"¿Está seguro eliminar todos los Piezómetros?")
        dlg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        dlg.setIcon(QMessageBox.Question)
        result = dlg.exec()
        if result == QMessageBox.Yes:
            respuesta = PiezometroController.ctrlEliminarPiezometrosCuerda(idzona)
            if respuesta:
                delete = PiezometroController.ctrlEliminarDataPiezometrosCuerda(idproyecto, respuesta)
                if delete:
                    TreeCheckbox.eliminarCheckboxGrupo(treewidget, idzona, grupo, tipo)
                    reiniciarvistas("Piezómetro")
                else:
                    mostrar_mensaje("Eliminar Piezómetros", "Error al eliminar data Piezómetros.", "advertencia")
            else:
                mostrar_mensaje("Eliminar Piezómetros", "No se pudo eliminar los Piezómetros.", "advertencia")
    
    def actualizarPiezometroCuerda(idproyecto, idcomponente, idinstrumento, treewidget, nombregrupo, tipogrupo, subgrupo, reiniciarvistas, vista="DATOS"):
        loader = QUiLoader()
        ui_file_path = resource_path("ui/editarpiezometrocuerda.ui")
        ui_file = loader.load(ui_file_path, None)
        # Configurar el cuadro de diálogo
        dialog = QDialog()
        dialog.setWindowTitle("Actualizar Piezómetro Cuerda Vibrante")
        layout = QVBoxLayout()
        layout.addWidget(ui_file)
        dialog.setLayout(layout)
        # Obtener elementos para interactuar
        comboComponente = dialog.findChild(QComboBox, "cb_lista_componentes")
        nombrePiezo = dialog.findChild(QLineEdit, "input_nombre")
        serieSensor = dialog.findChild(QLineEdit, "input_serie_sensor")
        comboFormula = dialog.findChild(QComboBox, "combo_formula")
        botonFormula = dialog.findChild(QPushButton, "btn_formula")
        cargarIcono(botonFormula, ListaIconos.ICONOS["calculadora"])
        nortePiezo = dialog.findChild(QDoubleSpinBox, "input_norte")
        estePiezo = dialog.findChild(QDoubleSpinBox, "input_este")
        instalacionPiezo = dialog.findChild(QDoubleSpinBox, "input_cota_instalacion")
        fundacionPiezo = dialog.findChild(QDoubleSpinBox, "input_cota_fundacion")
        factorCalibracion = dialog.findChild(QDoubleSpinBox, "input_factor_calibracion")
        correccionTempe = dialog.findChild(QDoubleSpinBox, "input_correccion_temperatura")
        lecturaInicial = dialog.findChild(QDoubleSpinBox, "input_frecuencia_inicial")
        temperaInicial = dialog.findChild(QDoubleSpinBox, "input_tempeartura_inicial")
        presionInicial = dialog.findChild(QDoubleSpinBox, "input_presion_inicial")
        unidadLectura = dialog.findChild(QComboBox, "combo_unidad_frecuencia")
        constanteA = dialog.findChild(QDoubleSpinBox, "input_constante_a")
        constanteB = dialog.findChild(QDoubleSpinBox, "input_constante_b")
        constanteC = dialog.findChild(QDoubleSpinBox, "input_constante_c")
        inclinacionPiezo = dialog.findChild(QSpinBox, "input_inclinacion")
        azimutPiezo = dialog.findChild(QSpinBox, "input_azimut")
        factorConversion = dialog.findChild(QDoubleSpinBox, "input_factor_conversion")
        comentarioPiezo = dialog.findChild(QTextEdit, "input_comentario")
        lblrespuesta = dialog.findChild(QLabel, "label_mensaje_error")
        botonguardar = dialog.findChild(QPushButton, "btn_guardar_nuevo")
        # cargar data componentes
        componentes = ProyectoController.ctrlObtenerComponentesProyecto(idproyecto)
        if componentes:
            for fila in componentes:
                comboComponente.addItem(str(fila[2]), fila[0])
        comboFormula.addItem("Sin Fórmula", 0)
        formulas = PiezometroController.ctrlTraerListaFormulas()
        if formulas:
            for fila in formulas:
                comboFormula.addItem(str(fila[1]), fila[0])
        # mostrar data Piezómetro Cuerda
        nombreactual = ""
        idpiezo = 0
        datapiezo = PiezometroController.ctrlObtenerInfoPiezometroCuerda(idinstrumento)
        if datapiezo:
            idpiezo = datapiezo[0]
            comboComponente.setCurrentIndex(comboComponente.findData(idcomponente))
            comboFormula.setCurrentIndex(comboFormula.findData(datapiezo[2]))
            nombrePiezo.setText(str(datapiezo[3]))
            nombreactual = str(datapiezo[3])
            serieSensor.setText(str(datapiezo[4]))
            estePiezo.setValue(datapiezo[5])
            nortePiezo.setValue(datapiezo[6])
            instalacionPiezo.setValue(datapiezo[7])
            fundacionPiezo.setValue(datapiezo[8])
            inclinacionPiezo.setValue(datapiezo[9])
            azimutPiezo.setValue(datapiezo[10])
            factorCalibracion.setValue(datapiezo[11])
            correccionTempe.setValue(datapiezo[12])
            lecturaInicial.setValue(datapiezo[13])
            temperaInicial.setValue(datapiezo[14])
            presionInicial.setValue(datapiezo[15])
            unidadLectura.setCurrentText(str(datapiezo[16]))
            constanteA.setValue(datapiezo[17])
            constanteB.setValue(datapiezo[18])
            constanteC.setValue(datapiezo[19])
            factorConversion.setValue(datapiezo[20])
            comentarioPiezo.setPlainText(str(datapiezo[21]))
        def actualizarPiezometro():
            componente = comboComponente.currentData()
            nombrezona = comboComponente.currentText()
            idformula = comboFormula.currentData()
            nombre = nombrePiezo.text()
            serie = serieSensor.text()
            norte = nortePiezo.value()
            este = estePiezo.value()
            instalacion = instalacionPiezo.value()
            fundacion = fundacionPiezo.value()
            inclinacion = inclinacionPiezo.value()
            azimut = azimutPiezo.value()
            calibracion = factorCalibracion.value()
            tempecorrec = correccionTempe.value()
            frecuenini = lecturaInicial.value()
            temperaini = temperaInicial.value()
            presionini = presionInicial.value()
            unidad = unidadLectura.currentText()
            variablea = constanteA.value()
            variableb = constanteB.value()
            variablec = constanteC.value()
            conversion = factorConversion.value()
            comentario = comentarioPiezo.toPlainText()
            if nombre != "":
                datos = (idformula, nombre, serie, este, norte, instalacion, fundacion, inclinacion, azimut, calibracion, tempecorrec, frecuenini, temperaini, presionini, unidad, variablea, variableb, variablec, conversion, comentario, idpiezo)
                data = (componente, nombre, idinstrumento)
                respuesta = PiezometroController.ctrlActualizarPiezometroCuerda(datos, data)
                if respuesta:
                    dialog.close()
                    if str(idcomponente) == str(componente):
                        TreeCheckbox.actualizarTextoCheckboxEquipo(treewidget, idcomponente, nombregrupo, subgrupo, nombreactual, nombre)
                    else:
                        # Eliminar Piezómetro
                        TreeCheckbox.eliminarCheckbox(treewidget, nombregrupo, idinstrumento, subgrupo)
                        # Crear piezometro cuerda en nuevo componente
                        piezocu = InterfazController.ctrlListarComponentePiezometroCuerda(idinstrumento)
                        if piezocu:
                            if vista == "PIEZOMETROS":
                                TreeCheckbox.crearNuevoGrupoCheckboxesDoble(treewidget, nombrezona, componente, idproyecto, nombregrupo, tipogrupo, piezocu, subgrupo)
                            else:
                                TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidget, nombrezona, componente, idproyecto, nombregrupo, tipogrupo, piezocu, subgrupo)
                    reiniciarvistas("Piezómetro")
                else:
                    lblrespuesta.setText("¡Error al actualizar los datos!")
                    lblrespuesta.setStyleSheet("color: red;")
            else:
                lblrespuesta.setText("¡Algunos datos están vacíos!")
                lblrespuesta.setStyleSheet("color: orange;")
        # Inicializar botones
        lblrespuesta.setText("")
        botonguardar.clicked.connect(actualizarPiezometro)
        # mostrar dialogo
        dialog.exec()
    
    def eliminar_piezocuerda(idproyecto, idinstrumento, nombrepiezo, nombregrupo, tipolista, treewidget, reiniciarvistas):
        dlg = QMessageBox()
        dlg.setWindowTitle("Eliminar P. Cuerda Vibrante")
        dlg.setText(f"¿Está seguro eliminar el Piezómetro '{nombrepiezo}'?")
        dlg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        dlg.setIcon(QMessageBox.Question)
        result = dlg.exec()
        if result == QMessageBox.Yes:
            respuesta = PiezometroController.ctrlEliminarCuerdaVibrante(idinstrumento)
            if respuesta:
                delete = PiezometroController.ctrlEliminarCuerdaVibranteData(idproyecto, respuesta)
                if delete:
                    TreeCheckbox.eliminarCheckbox(treewidget, nombregrupo, idinstrumento, tipolista)
                    reiniciarvistas("Piezómetro")
                else:
                    mostrar_mensaje("Eliminar Piezómetro", "Error al eliminar data del piezómetro.", "advertencia")
            else:
                mostrar_mensaje("Eliminar Piezómetro", "No se pudo eliminar el piezómetro.", "advertencia")
    
    def cambiar_componente_piezomanuales(idcomponente, idproyecto, treewidget, nombregrupo, tipogrupo, subgrupo, reiniciarvistas, vista="DATOS"):
        dialog = QDialog()
        dialog.setWindowTitle("Componente Piezómetros Manuales")
        layout = QFormLayout(dialog)
        # Campo componente
        label_titulo = QLabel("Componente:")
        combo_componente = QComboBox()
        label_mensaje = QLabel("")
        label_mensaje.setAlignment(Qt.AlignCenter)
        label_mensaje.setStyleSheet("QLabel { color: red; }")
        # cargar data componentes
        componentes = ProyectoController.ctrlObtenerComponentesProyecto(idproyecto)
        if componentes:
            for fila in componentes:
                combo_componente.addItem(str(fila[2]), fila[0])
            combo_componente.setCurrentIndex(combo_componente.findData(idcomponente))
        # Botones
        layout.addRow(label_titulo)
        layout.addRow(combo_componente)
        layout.addRow(label_mensaje)
        button_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        # Cambiar los textos a español
        button_box.button(QDialogButtonBox.Save).setText("Guardar")
        button_box.button(QDialogButtonBox.Cancel).setText("Cancelar")
        layout.addWidget(button_box)
        # Conectar los botones a las funciones correspondientes
        def actualizarDatos():
            componente = combo_componente.currentData()
            nombrezona = combo_componente.currentText()
            if str(idcomponente) == str(componente):
                dialog.reject()
            else:
                respuesta = PiezometroController.ctrlCambiarComponentePiezometrosManuales(idcomponente, componente)
                if respuesta:
                    dialog.reject()
                    # Eliminar cuerdas
                    TreeCheckbox.eliminarCheckboxGrupo(treewidget, idcomponente, nombregrupo, tipogrupo)
                    # Crear cuerdas en nuevo componente
                    if vista == "PIEZOMETROS":
                        TreeCheckbox.crearNuevoGrupoCheckboxesDoble(treewidget, nombrezona, componente, idproyecto, nombregrupo, tipogrupo, respuesta, subgrupo)
                    else:
                        TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidget, nombrezona, componente, idproyecto, nombregrupo, tipogrupo, respuesta, subgrupo)
                    reiniciarvistas("Piezómetro")
                else:
                    label_mensaje.setText("Error al cambiar de componente.")
        button_box.accepted.connect(actualizarDatos)
        button_box.rejected.connect(dialog.reject)
        # Mostrar el diálogo
        dialog.setLayout(layout)
        dialog.exec()
    
    def eliminar_piezomanuales(idproyecto, idzona, grupo, tipo, treewidget, reiniciarvistas):
        dlg = QMessageBox()
        dlg.setWindowTitle("Eliminar Piezómetros Manuales")
        dlg.setText(f"¿Está seguro eliminar todos los Piezómetros?")
        dlg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        dlg.setIcon(QMessageBox.Question)
        result = dlg.exec()
        if result == QMessageBox.Yes:
            respuesta = PiezometroController.ctrlEliminarPiezometrosManuales(idzona)
            if respuesta:
                delete = PiezometroController.ctrlEliminarDataPiezometrosManuales(idproyecto, respuesta)
                if delete:
                    TreeCheckbox.eliminarCheckboxGrupo(treewidget, idzona, grupo, tipo)
                    reiniciarvistas("Piezómetro")
                else:
                    mostrar_mensaje("Eliminar Piezómetros", "Error al eliminar data Piezómetros.", "advertencia")
            else:
                mostrar_mensaje("Eliminar Piezómetros", "No se pudo eliminar los Piezómetros.", "advertencia")
    
    def actualizarPiezometroManual(idproyecto, idcomponente, idinstrumento, treewidget, nombregrupo, tipogrupo, subgrupo, reiniciarvistas, vista="DATOS"):
        loader = QUiLoader()
        ui_file_path = resource_path("ui/editarpiezometromanual.ui")
        ui_file = loader.load(ui_file_path, None)
        # Configurar el cuadro de diálogo
        dialog = QDialog()
        dialog.setWindowTitle("Actualizar Piezómetro Manual")
        layout = QVBoxLayout()
        layout.addWidget(ui_file)
        dialog.setLayout(layout)
        # Obtener elementos para interactuar
        comboComponente = dialog.findChild(QComboBox, "cb_lista_componentes")
        nombrePiezo = dialog.findChild(QLineEdit, "input_nombre")
        codigoPiezo = dialog.findChild(QLineEdit, "input_codigo")
        nortePiezo = dialog.findChild(QDoubleSpinBox, "input_norte")
        estePiezo = dialog.findChild(QDoubleSpinBox, "input_este")
        elevacionPiezo = dialog.findChild(QDoubleSpinBox, "input_elevacion")
        fundacionPiezo = dialog.findChild(QDoubleSpinBox, "input_fundacion")
        stickupPiezo = dialog.findChild(QDoubleSpinBox, "input_stickup")
        inclinacionPiezo = dialog.findChild(QSpinBox, "input_inclinacion")
        azimutPiezo = dialog.findChild(QSpinBox, "input_azimut")
        comentarioPiezo = dialog.findChild(QTextEdit, "input_comentario")
        lblrespuesta = dialog.findChild(QLabel, "label_mensaje_error")
        botonguardar = dialog.findChild(QPushButton, "btn_aceptar_nuevo")
        # cargar data componentes
        componentes = ProyectoController.ctrlObtenerComponentesProyecto(idproyecto)
        if componentes:
            for fila in componentes:
                comboComponente.addItem(str(fila[2]), fila[0])
            comboComponente.setEnabled(True)
        # mostrar data Piezómetro Cuerda
        nombreactual = ""
        idpiezo = 0
        datapiezo = PiezometroController.ctrlObtenerInfoPiezometroManual(idinstrumento)
        if datapiezo:
            idpiezo = datapiezo[0]
            comboComponente.setCurrentIndex(comboComponente.findData(idcomponente))
            nombrePiezo.setText(str(datapiezo[2]))
            nombreactual = str(datapiezo[2])
            codigoPiezo.setText(str(datapiezo[3]))
            estePiezo.setValue(datapiezo[4])
            nortePiezo.setValue(datapiezo[5])
            elevacionPiezo.setValue(datapiezo[6])
            fundacionPiezo.setValue(datapiezo[7])
            stickupPiezo.setValue(datapiezo[10])
            inclinacionPiezo.setValue(datapiezo[8])
            azimutPiezo.setValue(datapiezo[9])
            comentarioPiezo.setPlainText(str(datapiezo[11]))
        def actualizarPiezometroManual():
            componente = comboComponente.currentData()
            nombrezona = comboComponente.currentText()
            nombre = nombrePiezo.text() 
            codigo = codigoPiezo.text()
            norte = nortePiezo.value()
            este = estePiezo.value()
            nivel = elevacionPiezo.value()
            fundacion = fundacionPiezo.value()
            stick = stickupPiezo.value()
            inclinacion = inclinacionPiezo.value()
            azimut = azimutPiezo.value()
            comentario = comentarioPiezo.toPlainText()
            if nombre != "":
                datos = (nombre, codigo, norte, este, nivel, fundacion, inclinacion, azimut, stick, comentario, idpiezo)
                data = (componente, nombre, idinstrumento)
                respuesta = PiezometroController.ctrlActualizarPiezometroManual(datos, data)
                if respuesta:
                    dialog.close()
                    if str(idcomponente) == str(componente):
                        TreeCheckbox.actualizarTextoCheckboxEquipo(treewidget, idcomponente, nombregrupo, subgrupo, nombreactual, nombre)
                    else:
                        # Eliminar Piezómetro
                        TreeCheckbox.eliminarCheckbox(treewidget, nombregrupo, idinstrumento, subgrupo)
                        # Crear piezometro manual en nuevo componente
                        piezocu = InterfazController.ctrlListarComponentePiezometroManual(idinstrumento)
                        if piezocu:
                            if vista == "PIEZOMETROS":
                                TreeCheckbox.crearNuevoGrupoCheckboxesDoble(treewidget, nombrezona, componente, idproyecto, nombregrupo, tipogrupo, piezocu, subgrupo)
                            else:
                                TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidget, nombrezona, componente, idproyecto, nombregrupo, tipogrupo, piezocu, subgrupo)
                    reiniciarvistas("Piezómetro")
                else:
                    lblrespuesta.setText("¡Error al guardar los datos!")
                    lblrespuesta.setStyleSheet("color: red;")
            else:
                lblrespuesta.setText("¡Algunos campos están vacíos!")
                lblrespuesta.setStyleSheet("color: red;")
        # Inicializar botones
        lblrespuesta.setText("")
        botonguardar.clicked.connect(actualizarPiezometroManual)
        # mostrar dialogo
        dialog.exec()
    
    def eliminar_piezomanual(idproyecto, idinstrumento, nombrepiezo, nombregrupo, tipolista, treewidget, reiniciarvistas):
        dlg = QMessageBox()
        dlg.setWindowTitle("Eliminar Piezómetro Manual")
        dlg.setText(f"¿Está seguro eliminar el Piezómetro '{nombrepiezo}'?")
        dlg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        dlg.setIcon(QMessageBox.Question)
        result = dlg.exec()
        if result == QMessageBox.Yes:
            respuesta = PiezometroController.ctrlEliminarManualPiezometro(idinstrumento)
            if respuesta:
                delete = PiezometroController.ctrlEliminarPiezometroManualData(idproyecto, respuesta)
                if delete:
                    TreeCheckbox.eliminarCheckbox(treewidget, nombregrupo, idinstrumento, tipolista)
                    reiniciarvistas("Piezómetro")
                else:
                    mostrar_mensaje("Eliminar Piezómetro", "Error al eliminar data del piezómetro.", "advertencia")
            else:
                mostrar_mensaje("Eliminar Piezómetro", "No se pudo eliminar el piezómetro.", "advertencia")
    
    def mostrarDialogoFechasPiezometros(treewidget, idproyecto, idcomponente, idinstrumento, nombrecompo, nombrepiezo, fechamarcada, tipo, graficarnuevafechaspiezometros):
        loaderLoading = QUiLoader()
        ui_file_path = resource_path("ui/arbolcheckbox.ui")
        ui_file = loaderLoading.load(ui_file_path, None)
        dialogo = QDialog()
        dialogo.setWindowTitle("Lista de fechas")
        layout = QVBoxLayout()
        layout.addWidget(ui_file)
        dialogo.setLayout(layout)
        # Obtener elementos para interactuar
        lbltitulo = dialogo.findChild(QLabel, "label_nombre")
        treefechas = dialogo.findChild(QTreeWidget, "tree_fechas")
        botonaceptar = dialogo.findChild(QPushButton, "btn_aceptar")
        lbltitulo.setText("FECHAS DEL PIEZÓMETRO")
        treefechas.setHeaderLabels([nombrecompo])
        listafechas = PiezometroController.ctrlListarFechasPiezometro(tipo, idcomponente, idinstrumento, idproyecto)
        if listafechas:
            parent = QTreeWidgetItem(treefechas)
            parent.setText(0, nombrepiezo)
            parent.setText(1, "1")
            if fechamarcada:
                parent.setCheckState(0, Qt.PartiallyChecked)
            else:
                parent.setCheckState(0, Qt.Unchecked)
            parent.setFlags(parent.flags() | Qt.ItemIsUserCheckable)
            parent.setExpanded(True)
            for fechas in listafechas:
                item = QTreeWidgetItem(parent)
                item.setText(0, fechas[0])
                item.setText(1, "fecha")
                item.setCheckState(0, Qt.Unchecked)
                item.setFlags(item.flags() | Qt.ItemIsUserCheckable | Qt.ItemIsSelectable)
                if fechamarcada:
                    if fechas[0] == fechamarcada:
                        item.setCheckState(0, Qt.Checked)
        def marcadoDesmarcadoCheckbox(parent_item, column):
            TreeCheckbox.validarMarcadoUnicoCheckbox(parent_item, column)
        def obtenerFechasMarcadas():
            fechaelegida = None
            parent = treefechas.topLevelItem(0)
            if parent:
                for i in range(parent.childCount()):
                    hijo = parent.child(i)
                    if hijo.checkState(0) == Qt.Checked:
                        fechaelegida = hijo.text(0)
            dialogo.close()
            if fechaelegida:
                if tipo == "Automatizado":
                    TreeCheckbox.actualizarFechasCheckboxEquipo(treewidget, idcomponente, "Piezómetros Cuerda Vibrante", "piezometrocuerda", nombrepiezo, fechaelegida)
                else:
                    TreeCheckbox.actualizarFechasCheckboxEquipo(treewidget, idcomponente, "Piezómetros Casagrande", "piezometromanual", nombrepiezo, fechaelegida)
                graficarnuevafechaspiezometros(tipo)
        # conectar funciones
        treefechas.itemClicked.connect(marcadoDesmarcadoCheckbox)
        botonaceptar.clicked.connect(obtenerFechasMarcadas)
        dialogo.exec()
    
    def cambiar_componente_bloque_cuerda(idproyecto, idcomponente, parent, treewidget, nombregrupo, tipogrupo, subgrupo, reiniciarvistas):
        dialog = QDialog()
        dialog.setWindowTitle("Mover Cuerda Vibrante Componente")
        layout = QFormLayout(dialog)
        # Campo componente
        label_titulo = QLabel("Componente:")
        combo_componente = QComboBox()
        label_mensaje = QLabel("")
        label_mensaje.setAlignment(Qt.AlignCenter)
        label_mensaje.setStyleSheet("QLabel { color: red; }")
        # cargar data componentes
        componentes = ProyectoController.ctrlObtenerComponentesProyecto(idproyecto)
        if componentes:
            for fila in componentes:
                combo_componente.addItem(str(fila[2]), fila[0])
            combo_componente.setCurrentIndex(combo_componente.findData(idcomponente))
        # Botones
        layout.addRow(label_titulo)
        layout.addRow(combo_componente)
        layout.addRow(label_mensaje)
        button_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        # Cambiar los textos a español
        button_box.button(QDialogButtonBox.Save).setText("Guardar")
        button_box.button(QDialogButtonBox.Cancel).setText("Cancelar")
        layout.addWidget(button_box)
        # Conectar los botones a las funciones correspondientes
        def actualizarDatos():
            componente = combo_componente.currentData()
            nombrezona = combo_componente.currentText()
            if str(idcomponente) == str(componente):
                dialog.reject()
            else:
                result = False
                hijos_marcados = []
                for i in range(parent.childCount()):
                    hijo = parent.child(i)
                    if hijo is not None and hijo.checkState(0) == Qt.Checked:
                        idinstrumento = hijo.text(2)
                        respuesta = PiezometroController.ctrlCambiarPiezometroComponente(idinstrumento, componente)
                        if respuesta:
                            hijos_marcados.append(hijo)
                            result = True
                if result:
                    dialog.reject()
                    for hijo in hijos_marcados:
                        idinstrumento = hijo.text(2)
                        TreeCheckbox.eliminarCheckbox(treewidget, nombregrupo, idinstrumento, subgrupo)
                        # Crear prisma en nuevo componente
                        piezometro = InterfazController.ctrlListarComponentePiezometroCuerda(idinstrumento)
                        if piezometro:
                            TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidget, nombrezona, componente, idproyecto, nombregrupo, tipogrupo, piezometro, subgrupo)
                    reiniciarvistas("Piezómetros Cuerda Vibrante")
                    # Limpiar tabla
                    from views.datos_view import DatosView
                    from modules.datos.vistaDatos import VistaDatos
                    tabla =  DatosView.main.findChild(QTableView, "table_datos")
                    VistaDatos.limpiarTablaDatos(tabla)
                else:
                    label_mensaje.setText("Error al cambiar de componente.")
        button_box.accepted.connect(actualizarDatos)
        button_box.rejected.connect(dialog.reject)
        # Mostrar el diálogo
        dialog.setLayout(layout)
        dialog.exec()
    
    def cambiar_componente_bloque_casagrande(idproyecto, idcomponente, parent, treewidget, nombregrupo, tipogrupo, subgrupo, reiniciarvistas):
        dialog = QDialog()
        dialog.setWindowTitle("Mover Casagrande Componente")
        layout = QFormLayout(dialog)
        # Campo componente
        label_titulo = QLabel("Componente:")
        combo_componente = QComboBox()
        label_mensaje = QLabel("")
        label_mensaje.setAlignment(Qt.AlignCenter)
        label_mensaje.setStyleSheet("QLabel { color: red; }")
        # cargar data componentes
        componentes = ProyectoController.ctrlObtenerComponentesProyecto(idproyecto)
        if componentes:
            for fila in componentes:
                combo_componente.addItem(str(fila[2]), fila[0])
            combo_componente.setCurrentIndex(combo_componente.findData(idcomponente))
        # Botones
        layout.addRow(label_titulo)
        layout.addRow(combo_componente)
        layout.addRow(label_mensaje)
        button_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        # Cambiar los textos a español
        button_box.button(QDialogButtonBox.Save).setText("Guardar")
        button_box.button(QDialogButtonBox.Cancel).setText("Cancelar")
        layout.addWidget(button_box)
        # Conectar los botones a las funciones correspondientes
        def actualizarDatos():
            componente = combo_componente.currentData()
            nombrezona = combo_componente.currentText()
            if str(idcomponente) == str(componente):
                dialog.reject()
            else:
                result = False
                hijos_marcados = []
                for i in range(parent.childCount()):
                    hijo = parent.child(i)
                    if hijo is not None and hijo.checkState(0) == Qt.Checked:
                        idinstrumento = hijo.text(2)
                        respuesta = PiezometroController.ctrlCambiarPiezometroComponente(idinstrumento, componente)
                        if respuesta:
                            hijos_marcados.append(hijo)
                            result = True
                if result:
                    dialog.reject()
                    for hijo in hijos_marcados:
                        idinstrumento = hijo.text(2)
                        TreeCheckbox.eliminarCheckbox(treewidget, nombregrupo, idinstrumento, subgrupo)
                        # Crear prisma en nuevo componente
                        piezometro = InterfazController.ctrlListarComponentePiezometroManual(idinstrumento)
                        if piezometro:
                            TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidget, nombrezona, componente, idproyecto, nombregrupo, tipogrupo, piezometro, subgrupo)
                    reiniciarvistas("Piezómetros Casagrande")
                    # Limpiar tabla
                    from views.datos_view import DatosView
                    from modules.datos.vistaDatos import VistaDatos
                    tabla =  DatosView.main.findChild(QTableView, "table_datos")
                    VistaDatos.limpiarTablaDatos(tabla)
                else:
                    label_mensaje.setText("Error al cambiar de componente.")
        button_box.accepted.connect(actualizarDatos)
        button_box.rejected.connect(dialog.reject)
        # Mostrar el diálogo
        dialog.setLayout(layout)
        dialog.exec()
    