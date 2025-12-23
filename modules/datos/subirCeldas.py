import pandas as pd
import re
from openpyxl import load_workbook
from PySide6.QtUiTools import QUiLoader
from PySide6.QtWidgets import (QDialog, QVBoxLayout, QComboBox, QHeaderView, QPushButton, QMenu, QTableWidget, QFormLayout,
                            QDialogButtonBox, QFileDialog, QMessageBox, QLabel, QDoubleSpinBox, QLineEdit, QTreeWidget, QTableView)
from PySide6.QtGui import QPen, QColor
from PySide6.QtCore import Qt
from datetime import datetime,time
from utils.common.rutasarchivos import resource_path
from utils.generic.cargariconos import cargarIcono
from utils.common.alertas import mostrar_mensaje
from utils.shared.pegarDatosTabla import configurar_tabla_para_pegado
from utils.shared.pegarDatosTabla import pegar_desde_portapapeles
from utils.shared.arbolmarcado import TreeCheckbox
from utils.common.metodosGenerales import MetodosGenerales
from controllers.ProyectoController import ProyectoController
from controllers.CeldaController import CeldaController
from controllers.InterfazController import InterfazController

# Pintar columna oculta de tabla
class CustomHeaderView(QHeaderView):
    def __init__(self, orientation, parent=None):
        super().__init__(orientation, parent)
    
    def paintSection(self, painter, rect, logicalIndex):
        super().paintSection(painter, rect, logicalIndex)
        
        # Verificar si la columna siguiente está oculta para dibujar la línea doble
        table_widget = self.parent()
        if logicalIndex < table_widget.columnCount() - 1 and table_widget.isColumnHidden(logicalIndex + 1):
            painter.save()
            pen = QPen(QColor(255, 0, 0), 2, Qt.SolidLine)  # Cambia el color a rojo
            painter.setPen(pen)

            # Dibujar la primera línea
            painter.drawLine(rect.right() - 2, rect.top(), rect.right() - 2, rect.bottom())

            # Dibujar la segunda línea para el efecto de línea doble
            painter.drawLine(rect.right() - 6, rect.top(), rect.right() - 6, rect.bottom())
            painter.restore()
            
class SubirCeldas:
    
    def registrarDataCeldas(main, proyectoid):
        loader = QUiLoader()
        ui_file_path = resource_path("ui/cargardataceldas.ui")
        dialogo_data_celda = loader.load(ui_file_path, None)
        # Configurar el cuadro de diálogo principal
        dialog_data_celda = QDialog()
        dialog_data_celda.setWindowTitle("Celdas de Asentamiento")
        layout_registrar_data = QVBoxLayout()
        layout_registrar_data.addWidget(dialogo_data_celda)
        dialog_data_celda.setLayout(layout_registrar_data)
        combo_celdas_asentamiento = dialog_data_celda.findChild(QComboBox, "combo_celdas")
        tabladata = dialog_data_celda.findChild(QTableWidget, "table_celdas_calculo")
        boton_guardar = dialog_data_celda.findChild(QPushButton, "btn_guardar_data_celda")
        lblrespuesta = dialog_data_celda.findChild(QLabel, "label_mensaje_estado")
        # Verifica si hay datos
        datos = CeldaController.ctrlObtenerCeldasAsentamiento(proyectoid)
        if datos:
            for id_celda, nombre_celda in datos:
                combo_celdas_asentamiento.addItem(nombre_celda, id_celda)
        else:
            combo_celdas_asentamiento.addItem("Sin celdas")
            boton_guardar.setEnabled(False)
        # agregar una celda
        row_position = tabladata.rowCount()
        tabladata.insertRow(row_position)
        # Configurar el menú contextual en la cabecera
        custom_header = CustomHeaderView(Qt.Horizontal, tabladata)
        tabladata.setHorizontalHeader(custom_header)
        header = tabladata.horizontalHeader()
        header.setContextMenuPolicy(Qt.CustomContextMenu)
        def mostrarMenuCabecera(pos):
            menu = QMenu()
            columna = header.logicalIndexAt(pos)
            if tabladata.isColumnHidden(columna):
                toggle_action = menu.addAction(f"Mostrar Columna '{tabladata.horizontalHeaderItem(columna).text()}'")
                toggle_action.triggered.connect(lambda: tabladata.showColumn(columna))
            else:
                toggle_action = menu.addAction(f"Ocultar Columna '{tabladata.horizontalHeaderItem(columna).text()}'")
                toggle_action.triggered.connect(lambda: tabladata.hideColumn(columna))
            if tabladata.isColumnHidden(columna + 1):
                # Buscar todas las columnas ocultas consecutivas y agregarlas al menú
                hidden_columns = []
                next_index = columna + 1
                while next_index < tabladata.columnCount() and tabladata.isColumnHidden(next_index):
                    hidden_columns.append(next_index)
                    next_index += 1
                # Agregar cada columna oculta consecutiva al menú
                for col in hidden_columns:
                    show_action = menu.addAction(f"Mostrar Columna '{tabladata.horizontalHeaderItem(col).text()}'")
                    show_action.triggered.connect(lambda checked, col=col: tabladata.showColumn(col))
            # Opción para mostrar todas las columnas
            show_all_action = menu.addAction("Mostrar Todas las Columnas")
            show_all_action.triggered.connect(lambda: [tabladata.showColumn(col) for col in range(tabladata.columnCount())])
            # Mostrar menu
            menu.exec(header.mapToGlobal(pos))
        def menuPegadoDataTabla(pos):
            context_menu = QMenu()
            paste_action = context_menu.addAction("Pegar aquí")
            paste_action.triggered.connect(lambda:  pegar_desde_portapapeles(tabladata))
            context_menu.exec(tabladata.mapToGlobal(pos))
        def registrarDataCeldaAsentamiento():
            id_celda = combo_celdas_asentamiento.currentData()
            filas = tabladata.rowCount()
            if filas > 0 and proyectoid != 0:
                data = []
                estado = False
                for row in range(filas):
                    datosfila = []
                    datosfila.append(id_celda)
                    fila_valida = True
                    c = 0
                    for column in range(tabladata.columnCount()):
                        item = tabladata.item(row, column)
                        valor = item.text().strip() if item else ""
                        if valor != "":
                            if column == 0:
                                valor = MetodosGenerales.validarFormatoFecha(valor)
                                if not valor:
                                    mensaje = "La fecha no tiene un formato adecuado."
                                    fila_valida = False
                                    break
                            elif column == 1:
                                valor = MetodosGenerales.validarFormatoHora(valor)
                                if not valor:
                                    mensaje = "La hora no tiene un formato adecuado."
                                    fila_valida = False
                                    break
                            elif column > 1 and column < 6:
                                if not MetodosGenerales.validarEsNumero(valor):
                                    mensaje = "Algunas lecturas no son numéricas."
                                    fila_valida = False
                                    break
                        else:
                            if column == 0:
                                fila_valida = False
                                mensaje = "La fecha está vacía."
                            elif column == 1:
                                valor = "00:00:00"
                            elif column == 2:
                                valor = 0
                            elif column == 3:
                                valor = 0
                            elif column == 4:
                                valor = 0
                            elif column == 5:
                                fila_valida = False
                                mensaje = "La medida (mca) está vacía."
                            c += 1
                        datosfila.append(valor)
                    if c != 7:
                        if fila_valida and len(datosfila) == 8:
                            data.append(datosfila)
                            estado = True
                        else:
                            estado = False
                            break
                if estado:
                    respuesta = CeldaController.ctrlRegistrarDataCelda(proyectoid, data)
                    if respuesta:
                        lblrespuesta.setText("Registrado Correctamente")
                        lblrespuesta.setStyleSheet("color: green;")
                        # Limpiar filas tabla
                        tabladata.setRowCount(0)
                        tabladata.insertRow(0)
                        # actualizar árbol checkbox
                        data = CeldaController.ctrlTraerDataCeldaAsentamiento(id_celda)
                        if data:
                            idinstrumento, idcomponente, nombrezona = data[0], data[1], data[2]
                            treewidgetdatos = main.findChild(QTreeWidget, "tree_actual_datos")
                            treewidgetvisor = main.findChild(QTreeWidget, "tree_actual_visor")
                            treewidgetceldas = main.findChild(QTreeWidget, "tree_actual_celdas")
                            TreeCheckbox.eliminarCheckbox(treewidgetdatos, "Celdas de Asentamiento", idinstrumento, "celda")
                            TreeCheckbox.eliminarCheckbox(treewidgetvisor, "Celdas de Asentamiento", idinstrumento, "celda")
                            TreeCheckbox.eliminarCheckbox(treewidgetceldas, "Celdas de Asentamiento", idinstrumento, "celda")
                            # Crear piezometro cuerda en nuevo componente
                            celda = InterfazController.ctrlListarComponenteCelda(idinstrumento)
                            if celda:
                                TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidgetdatos, nombrezona, idcomponente, proyectoid, "Celdas de Asentamiento", "7", celda, "celda")
                                TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidgetvisor, nombrezona, idcomponente, proyectoid, "Celdas de Asentamiento", "7", celda, "celda")
                                TreeCheckbox.crearNuevoGrupoCheckboxesDoble(treewidgetceldas, nombrezona, idcomponente, proyectoid, "Celdas de Asentamiento", "1", celda, "celda")
                    else:
                        lblrespuesta.setText("Error al regitrar.")
                        lblrespuesta.setStyleSheet("color: red;")
                else:
                    lblrespuesta.setText(f"En la fila {len(data) + 1}: {mensaje}")
                    lblrespuesta.setStyleSheet("color: orange;")
                    
        # conectar señales
        header.customContextMenuRequested.connect(mostrarMenuCabecera)
        tabladata.setContextMenuPolicy(Qt.CustomContextMenu)
        tabladata.customContextMenuRequested.connect(menuPegadoDataTabla)
        configurar_tabla_para_pegado(tabladata)
        boton_guardar.clicked.connect(registrarDataCeldaAsentamiento)
        dialog_data_celda.exec()
    
    def cargarDataFormatosCeldas(main, proyectoid, tipo):
        loader = QUiLoader()
        ui_file_path = resource_path("ui/cargardataformato.ui")
        ui_file = loader.load(ui_file_path, None)
        dialogo = QDialog()
        dialogo.setWindowTitle("Data Celdas")
        layout_procesar_data = QVBoxLayout()
        layout_procesar_data.addWidget(ui_file)
        dialogo.setLayout(layout_procesar_data)
        ruta = "resources/iconos/fontawesome/solid/file-arrow-up.svg"
        botonSubir = dialogo.findChild(QPushButton, "btn_cargar_archivo")
        cargarIcono(botonSubir, ruta)
        ubicacion_archivo = dialogo.findChild(QLineEdit, "input_archivo")
        ubicacion_archivo.setReadOnly(True)
        comboComponentes = dialogo.findChild(QComboBox, "combo_componentes")
        labelRespuesta = dialogo.findChild(QLabel, "label_mensaje")
        botonAceptar = dialogo.findChild(QPushButton, "btn_aceptar")
        componentes = ProyectoController.ctrlObtenerComponentesProyecto(proyectoid)
        if len(componentes) > 0:
            for fila in componentes:
                comboComponentes.addItem(str(fila[2]), fila[0])
            comboComponentes.setEnabled(True)
        else:
            comboComponentes.addItem("Sin Componentes")
            comboComponentes.setEnabled(False)
            botonAceptar.setEnabled(False)
        def cargar_archivo():
            file_names, _ = QFileDialog.getOpenFileNames(None, "Cargar Archivos", "", "Archivos Excel (*.xlsx)")
            if file_names:
                ubicacion_archivo.setText("\n".join(file_names))
        def procesar_archivo():
            if not ubicacion_archivo.text().strip():
                labelRespuesta.setText("No se cargó ningún archivo.")
                labelRespuesta.setStyleSheet("color: red;")
                return
            else:
                try:
                    idcompo = comboComponentes.currentData()
                    if tipo == "FORMATO":
                        respuesta, equipos, erroneos = SubirCeldas.registrarFormatoDataCeldas(proyectoid, ubicacion_archivo.text(), idcompo)
                    else:
                        respuesta, equipos, erroneos = SubirCeldas.registrarDataExcelCeldas(proyectoid, ubicacion_archivo.text(), idcompo)
                    if respuesta:
                        ubicacion_archivo.clear()
                        if len(erroneos) > 0:
                            labelRespuesta.setText(f"Archivos erróneos: {erroneos}")
                        else:
                            labelRespuesta.setText("Guardado correctamente.")
                        labelRespuesta.setStyleSheet("color: green;")
                        # actualizar árbol checkbox
                        for idcelda in equipos:
                            data = CeldaController.ctrlTraerDataCeldaAsentamiento(idcelda)
                            if data:
                                idinstrumento, idcomponente, nombrezona = data[0], data[1], data[2]
                                treewidgetdatos = main.findChild(QTreeWidget, "tree_actual_datos")
                                treewidgetvisor = main.findChild(QTreeWidget, "tree_actual_visor")
                                treewidgetceldas = main.findChild(QTreeWidget, "tree_actual_celdas")
                                TreeCheckbox.eliminarCheckbox(treewidgetdatos, "Celdas de Asentamiento", idinstrumento, "celda")
                                TreeCheckbox.eliminarCheckbox(treewidgetvisor, "Celdas de Asentamiento", idinstrumento, "celda")
                                TreeCheckbox.eliminarCheckbox(treewidgetceldas, "Celdas de Asentamiento", idinstrumento, "celda")
                                # Crear piezometro cuerda en nuevo componente
                                celda = InterfazController.ctrlListarComponenteCelda(idinstrumento)
                                if celda:
                                    TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidgetdatos, nombrezona, idcomponente, proyectoid, "Celdas de Asentamiento", "7", celda, "celda")
                                    TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidgetvisor, nombrezona, idcomponente, proyectoid, "Celdas de Asentamiento", "7", celda, "celda")
                                    TreeCheckbox.crearNuevoGrupoCheckboxesDoble(treewidgetceldas, nombrezona, idcomponente, proyectoid, "Celdas de Asentamiento", "1", celda, "celda")
                    else:
                        if len(erroneos) > 0:
                            labelRespuesta.setText(f"Error en los archivos: {erroneos}")
                        else:
                            labelRespuesta.setText("No se guardó la data.")
                        labelRespuesta.setStyleSheet("color: red;")
                except ValueError as e:
                    labelRespuesta.setText(str(e))
                    labelRespuesta.setStyleSheet("color: red;")
        botonSubir.clicked.connect(cargar_archivo)
        botonAceptar.clicked.connect(procesar_archivo)
        dialogo.exec()
    
    def registrarFormatoDataCeldas(proyectoid, ubicacion, idcomponente):
        erroneos = []
        data = []
        equipos = []
        respuesta = False
        encabezado = ['Fecha', 'Hora', 'Frecuencia (Digits)', 'Frecuencia (Hz)', 'Temperatura (°C)', 'Desplazamiento (m)', 'Observación']
        archivos = ubicacion.split("\n")
        for file_name in archivos:
            file_name = file_name.strip()
            if not file_name or not file_name.endswith('.xlsx'):
                continue
            try:
                df_header = pd.read_excel(file_name, header=None, nrows=1, skiprows=14, engine='openpyxl')
                encabezados_archivo = [str(col).strip() for col in df_header.iloc[0, :len(encabezado)]]
                if encabezados_archivo != encabezado:
                    erroneos.append(file_name.split("/")[-1])
                    continue
                wb = load_workbook(file_name, data_only=True)
                hoja = wb.active
                nombrecelda = hoja["B9"].value
                marcacelda = hoja["B10"].value
                modelocelda = hoja["B11"].value
                seriecelda = hoja["B12"].value
                cf = hoja["B13"].value
                tk = hoja["B14"].value
                instalacion = hoja["F9"].value
                fundacion = hoja["F10"].value
                coordeste = hoja["F11"].value
                coordnorte = hoja["F12"].value
                superficie = hoja["F13"].value
                rangocelda = hoja["F14"].value
                wb.close()
                if pd.isna(nombrecelda) or proyectoid == 0 or not idcomponente:
                    erroneos.append(file_name.split("/")[-1])
                    continue
                idcelda = None
                respu, info = CeldaController.ctrlComprobarExisteNombreCelda(proyectoid, nombrecelda)
                if respu:
                    idcelda = info[0]
                    celda_data = {
                        "marca_celda": marcacelda,
                        "modelo_celda": modelocelda,
                        "rango_celda": rangocelda,
                        "cf_celda": cf,
                        "cota_fundacion_celda": fundacion,
                        "coordenada_este_celda": coordeste,
                        "coordenada_norte_celda": coordnorte,
                        "cota_instalacion_celda": instalacion,
                        "tk_celda": tk,
                        "idcelda": idcelda,
                    }
                    if float(coordeste)!=0 and float(coordnorte)!=0:
                        # Llamar al método para guardar en la base de datos
                        rpt = CeldaController.ctrlActualizarCeldaExcel(celda_data)
                else:
                    if pd.isna(superficie):
                        continue
                    if pd.isna(coordnorte):
                        coordnorte = 0
                    else:
                        try:
                            coordnorte = float(coordnorte)
                        except ValueError:
                            coordnorte = 0
                    if pd.isna(coordeste):
                        coordeste = 0
                    else:
                        try:
                            coordeste = float(coordeste)
                        except ValueError:
                            coordeste = 0
                    if pd.isna(instalacion):
                        instalacion = 0
                    else:
                        try:
                            instalacion = float(instalacion)
                        except ValueError:
                            instalacion = 0
                    if pd.isna(fundacion):
                        fundacion = 0
                    else:
                        try:
                            fundacion = float(fundacion)
                        except ValueError:
                            fundacion = 0
                    if pd.isna(cf):
                        cf = 0
                    else:
                        try:
                            cf = float(cf)
                        except ValueError:
                            cf = 0
                    if pd.isna(tk):
                        tk = 0
                    else:
                        try:
                            tk = float(tk)
                        except ValueError:
                            tk = 0
                    datoscelda = {
                        "proyecto": proyectoid,
                        "nombre_celda": nombrecelda,
                        "marca_celda": marcacelda,
                        "modelo_celda": modelocelda,
                        "rango_celda": rangocelda,
                        "frecuencia_inicial": 0,
                        "cf_celda": cf,
                        "cota_superficie_celda": superficie,
                        "cota_fundacion_celda": fundacion,
                        "coordenada_este_celda": coordeste,
                        "coordenada_norte_celda": coordnorte,
                        "cota_instalacion_celda": instalacion,
                        "temperatura_inicial_celda": 0,
                        "tk_celda": tk,
                    }
                    respues = CeldaController.ctrlRegistrarCeldaFormato(idcomponente, datoscelda)
                    if respues:
                        idcelda = respues
                if idcelda is not None:
                    df = pd.read_excel(file_name, header=None, skiprows=15, engine='openpyxl')
                    df.columns = ['fecha', 'hora', 'frecuen', 'frecuencia', 'temperatura', 'desplaza', 'observacion']
                    for _, row in df.iterrows():
                        fecha = row['fecha']
                        hora = row['hora']
                        desplaza = row['desplaza']
                        observacion = row['observacion']
                        if pd.isna(fecha) or pd.isna(desplaza):
                            continue
                        # Manejo de la columna 'fecha'
                        if isinstance(fecha, (pd.Timestamp, datetime)):
                            fecha = fecha.date()  # Convertir a date
                            fecha = fecha.strftime('%Y-%m-%d')
                        elif isinstance(fecha, str):
                            fecha = MetodosGenerales.validarFormatoFecha(fecha)
                            if fecha is None:
                                continue
                        else:
                            continue
                        # Manejo de la columna 'hora'
                        if isinstance(hora, (pd.Timestamp, datetime)):
                            hora = hora.time()
                            hora = hora.strftime('%H:%M:%S')
                        elif isinstance(hora, time):
                            hora = hora.strftime('%H:%M:%S')
                        elif isinstance(hora, str):
                            hora = MetodosGenerales.validarFormatoHora(hora) or "00:00:00"
                        else:
                            hora = "00:00:00"
                        try:
                            desplaza = float(desplaza)
                        except (ValueError, TypeError):
                            continue
                        frecuen = float(row['frecuen']) if not pd.isna(row['frecuen']) else 0
                        frecu = float(row['frecuencia']) if not pd.isna(row['frecuencia']) else 0
                        tempe = float(row['temperatura']) if not pd.isna(row['temperatura']) else 0
                        tempe = float(row['temperatura']) if not pd.isna(row['temperatura']) else 0
                        observa = observacion if not pd.isna(observacion) else ""
                        data.append((idcelda, fecha, hora, frecuen, frecu, tempe, desplaza, observa))
                    if data:
                        respon = CeldaController.ctrlRegistrarDataCelda(proyectoid, data)
                        if respon:
                            equipos.append(idcelda)
                            respuesta = True
                        else:
                            erroneos.append(file_name.split("/")[-1])
                    else:
                        erroneos.append(file_name.split("/")[-1])
                else:
                    erroneos.append(file_name.split("/")[-1])
            except Exception:
                erroneos.append(file_name.split("/")[-1])
        return respuesta, equipos, erroneos
    
    def registrarDataExcelCeldas(proyectoid, ubicacion, idcomponente):
        erroneos = []
        data = []
        equipos = []
        respuesta = False
        # Definir encabezado esperado y las celdas donde debe estar cada columna
        encabezado_esperado = {
            'A12': 'Fecha',
            'D12': 'Frecuencia (Digits)',
            'I12': 'Frecuencia (Hz)',
            'M12': 'Temperatura (°C)',
            'R12': 'Desplazamiento (m)',
            'Y12': 'Cota (m s.n.m.)',
            'AD12': 'Observación'
        }
        archivos = ubicacion.split("\n")
        for file_name in archivos:
            file_name = file_name.strip()
            if not file_name or not file_name.endswith('.xlsx'):
                continue
            try:
                # Obtener todas las hojas del archivo Excel
                wb = load_workbook(file_name, data_only=True)
                sheet_names = wb.sheetnames
                # Procesar cada hoja
                for sheet_name in sheet_names:
                    try:
                        # Seleccionar la hoja actual
                        hoja = wb[sheet_name]
                        # Validar el encabezado leyendo celdas específicas
                        encabezado_valido = True
                        for celda, valor_esperado in encabezado_esperado.items():
                            valor_celda = hoja[celda].value
                            if valor_celda is None:
                                valor_celda = ""
                            valor_celda = str(valor_celda).strip()
                            if valor_celda != valor_esperado:
                                encabezado_valido = False
                                break
                        if not encabezado_valido:
                            continue
                        # Obtener data general de la hoja actual
                        nombrecelda = hoja["C6"].value
                        marcacelda = hoja["C7"].value
                        modelocelda = hoja["C8"].value
                        seriecelda = hoja["C9"].value
                        rangocelda = hoja["C10"].value
                        instalacion = hoja["P7"].value
                        fundacion = hoja["P7"].value
                        coordeste = hoja["N9"].value
                        coordnorte = hoja["N10"].value
                        superficie = hoja["P6"].value
                        cf = hoja["W10"].value
                        tk = hoja["AE10"].value
                        frecuenini = hoja["Y5"].value
                        temperaini = hoja["AD5"].value
                        # Validar datos
                        if pd.isna(nombrecelda) or proyectoid == 0 or not idcomponente:
                            continue  # Continuar con la siguiente hoja
                        idcelda = None
                        respu, info = CeldaController.ctrlComprobarExisteNombreCelda(proyectoid, nombrecelda)
                        if respu:
                            idcelda = info[0]
                        else:
                            if pd.isna(superficie):
                                continue
                            if pd.isna(coordnorte):
                                coordnorte = 0
                            else:
                                try:
                                    coordnorte = float(coordnorte)
                                except ValueError:
                                    coordnorte = 0
                            if pd.isna(coordeste):
                                coordeste = 0
                            else:
                                try:
                                    coordeste = float(coordeste)
                                except ValueError:
                                    coordeste = 0
                            if pd.isna(instalacion):
                                instalacion = 0
                            else:
                                try:
                                    instalacion = float(instalacion)
                                except ValueError:
                                    instalacion = 0
                            if pd.isna(fundacion):
                                fundacion = 0
                            else:
                                try:
                                    fundacion = float(fundacion)
                                except ValueError:
                                    fundacion = 0
                            if pd.isna(cf):
                                cf = 0
                            else:
                                try:
                                    cf = float(cf)
                                except ValueError:
                                    cf = 0
                            if pd.isna(tk):
                                tk = 0
                            else:
                                try:
                                    tk = float(tk)
                                except ValueError:
                                    tk = 0
                            coincidencia = re.search(r'-?\d+(?:\.\d+)?', str(rangocelda).strip() if rangocelda else '')
                            numerorango = float(coincidencia.group()) if coincidencia else 0
                            if pd.isna(frecuenini):
                                frecuenini = 0
                            else:
                                try:
                                    frecuenini = float(frecuenini)
                                except ValueError:
                                    frecuenini = 0
                            if pd.isna(temperaini):
                                temperaini = 0
                            else:
                                try:
                                    temperaini = float(temperaini)
                                except ValueError:
                                    temperaini = 0
                            datoscelda = {
                                "proyecto": proyectoid,
                                "nombre_celda": nombrecelda,
                                "marca_celda": marcacelda,
                                "modelo_celda": modelocelda,
                                "serie_celda": seriecelda,
                                "rango_celda": numerorango,
                                "frecuencia_inicial": frecuenini,
                                "cf_celda": cf,
                                "cota_superficie_celda": superficie,
                                "cota_fundacion_celda": fundacion,
                                "coordenada_este_celda": coordeste,
                                "coordenada_norte_celda": coordnorte,
                                "cota_instalacion_celda": instalacion,
                                "temperatura_inicial_celda": temperaini,
                                "tk_celda": tk,
                            }
                            respues = CeldaController.ctrlRegistrarCeldaFormato(idcomponente, datoscelda)
                            if respues:
                                idcelda = respues
                                equipos.append(idcelda)
                        if idcelda is not None:
                            # Leer datos desde la fila 13 en adelante usando las columnas específicas
                            sheet_data = []
                            fila = 13  # Empezar desde la fila 13 (debajo del encabezado)
                            while True:
                                # Leer datos de la fila actual
                                fecha_valor = hoja[f'A{fila}'].value
                                digits_valor = hoja[f'D{fila}'].value
                                frecuencia_valor = hoja[f'I{fila}'].value
                                temperatura_valor = hoja[f'M{fila}'].value
                                desplaza_valor = hoja[f'R{fila}'].value
                                cota_valor = hoja[f'Y{fila}'].value
                                observacion_valor = hoja[f'AD{fila}'].value
                                # Si no hay fecha o mca, terminar la lectura
                                if pd.isna(fecha_valor) or pd.isna(desplaza_valor):
                                    # Verificar si es una fila completamente vacía
                                    if all(pd.isna(val) for val in [fecha_valor, digits_valor, frecuencia_valor, temperatura_valor, desplaza_valor, cota_valor, observacion_valor]):
                                        break
                                    # Si solo faltan fecha o mca, saltar esta fila
                                    fila += 1
                                    continue
                                # Manejo de la columna 'fecha'
                                if isinstance(fecha_valor, (pd.Timestamp, datetime)):
                                    fecha_procesada = fecha_valor.date()
                                    fecha_procesada = fecha_procesada.strftime('%Y-%m-%d')
                                elif isinstance(fecha_valor, str):
                                    fecha_procesada = MetodosGenerales.validarFormatoFecha(fecha_valor)
                                    if fecha_procesada is None:
                                        fila += 1
                                        continue
                                else:
                                    fila += 1
                                    continue
                                # Manejo de la columna 'hora'
                                hora = "00:00:00"
                                try:
                                    desplaza_procesada = float(desplaza_valor)
                                except (ValueError, TypeError):
                                    fila += 1
                                    continue
                                # Procesar otros valores numéricos
                                digits = float(digits_valor) if not pd.isna(digits_valor) else 0
                                frecu = float(frecuencia_valor) if not pd.isna(frecuencia_valor) else 0
                                tempe = float(temperatura_valor) if not pd.isna(temperatura_valor) else 0
                                # Procesar observación
                                observa = observacion_valor if not pd.isna(observacion_valor) else ""
                                sheet_data.append((idcelda, fecha_procesada, hora, digits, frecu, tempe, desplaza_procesada, observa))
                                fila += 1
                            # Agregar datos de esta hoja al conjunto total
                            if sheet_data:
                                data.extend(sheet_data)
                    except Exception as e:
                        continue
                # Cerrar el workbook después de procesar todas las hojas
                wb.close()
                # Procesar todos los datos del archivo si hay datos válidos
                if data:
                    respon = CeldaController.ctrlRegistrarDataCelda(proyectoid, data)
                    if respon:
                        respuesta = True
                    else:
                        erroneos.append(file_name.split("/")[-1])
                else:
                    erroneos.append(file_name.split("/")[-1])
            except Exception as e:
                erroneos.append(file_name.split("/")[-1])
        return respuesta, equipos, erroneos
    
    def cambiar_componente_celdas(idcomponente, idproyecto, treewidget, nombregrupo, tipogrupo, subgrupo, reiniciarvistas, vista="DATOS"):
        dialog = QDialog()
        dialog.setWindowTitle("Componente Celdas")
        layout = QFormLayout(dialog)
        # Campo componente
        label_titulo = QLabel("Componente:")
        combo_componente = QComboBox()
        label_mensaje = QLabel("")
        label_mensaje.setAlignment(Qt.AlignCenter)
        label_mensaje.setStyleSheet("QLabel { color: red; }")
        # cargar data componentes
        componentes = ProyectoController.ctrlObtenerComponentesProyecto(idproyecto)
        if componentes:
            for fila in componentes:
                combo_componente.addItem(str(fila[2]), fila[0])
            combo_componente.setCurrentIndex(combo_componente.findData(idcomponente))
        # Botones
        layout.addRow(label_titulo)
        layout.addRow(combo_componente)
        layout.addRow(label_mensaje)
        button_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        # Cambiar los textos a español
        button_box.button(QDialogButtonBox.Save).setText("Guardar")
        button_box.button(QDialogButtonBox.Cancel).setText("Cancelar")
        layout.addWidget(button_box)
        # Conectar los botones a las funciones correspondientes
        def actualizarDatos():
            componente = combo_componente.currentData()
            nombrezona = combo_componente.currentText()
            if str(idcomponente) == str(componente):
                dialog.reject()
            else:
                respuesta = CeldaController.ctrlCambiarComponenteCeldas(idcomponente, componente)
                if respuesta:
                    dialog.reject()
                    # Eliminar terreno
                    TreeCheckbox.eliminarCheckboxGrupo(treewidget, idcomponente, nombregrupo, tipogrupo)
                    # Crear pluvio en nuevo componente
                    if vista == "CELDAS":
                        TreeCheckbox.crearNuevoGrupoCheckboxesDoble(treewidget, nombrezona, componente, idproyecto, nombregrupo, tipogrupo, respuesta, subgrupo)
                    else:
                        TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidget, nombrezona, componente, idproyecto, nombregrupo, tipogrupo, respuesta, subgrupo)
                    reiniciarvistas("Celda")
                else:
                    label_mensaje.setText("Error al cambiar de componente.")
        button_box.accepted.connect(actualizarDatos)
        button_box.rejected.connect(dialog.reject)
        # Mostrar el diálogo
        dialog.setLayout(layout)
        dialog.exec()
    
    def eliminar_celdas(idproyecto, idzona, grupo, tipo, treewidget, reiniciarvistas):
        dlg = QMessageBox()
        dlg.setWindowTitle("Eliminar Celdas")
        dlg.setText(f"¿Está seguro eliminar todos las Celdas?")
        dlg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        dlg.setIcon(QMessageBox.Question)
        result = dlg.exec()
        if result == QMessageBox.Yes:
            respuesta = CeldaController.ctrlEliminarCeldas(idzona)
            if respuesta:
                delete = CeldaController.ctrlEliminarDataCeldas(idproyecto, respuesta)
                if delete:
                    TreeCheckbox.eliminarCheckboxGrupo(treewidget, idzona, grupo, tipo)
                    reiniciarvistas("Celda")
                else:
                    mostrar_mensaje("Eliminar Celdas", "Error al eliminar data Celdas.", "advertencia")
            else:
                mostrar_mensaje("Eliminar Celdas", "No se pudo eliminar las Celdas.", "advertencia")
    
    def actualizarCeldaAsentamiento(idproyecto, idcomponente, idinstrumento, treewidget, nombregrupo, tipogrupo, subgrupo, reiniciarvistas, vista="DATOS"):        
        loaderLoading = QUiLoader()        
        ui_file_path = resource_path("ui/registroceldas.ui")
        ui_file = loaderLoading.load(ui_file_path, None)
        dialogo = QDialog()
        dialogo.setWindowTitle("Actualizar Celda de Asentamiento")
        layout_procesar_data = QVBoxLayout()
        layout_procesar_data.addWidget(ui_file)
        dialogo.setLayout(layout_procesar_data)
        # tools
        comboComponente = dialogo.findChild(QComboBox, "combo_componente")
        nombre_celda = dialogo.findChild(QLineEdit, "nombre_celda")
        marca_celda = dialogo.findChild(QLineEdit, "marca_celda")
        modelo_celda = dialogo.findChild(QLineEdit, "modelo_celda")
        rango_celda = dialogo.findChild(QDoubleSpinBox, "dsb_rango_celda")
        frecuencia_inicial_celda = dialogo.findChild(QDoubleSpinBox, "dsb_frecuencia_inicial_celda")
        cf_celda = dialogo.findChild(QDoubleSpinBox, "dsb_cf_celda")
        cota_superficie_celda = dialogo.findChild(QDoubleSpinBox, "dsb_cota_superficie_celda")
        cota_fundacion_celda = dialogo.findChild(QDoubleSpinBox, "dsb_cota_fundacion_celda")
        coordenada_este_celda = dialogo.findChild(QDoubleSpinBox, "dsb_coordenada_este_celda")
        coordenada_norte_celda = dialogo.findChild(QDoubleSpinBox, "dsb_coordenada_norte_celda")
        cota_instalacion_celda = dialogo.findChild(QDoubleSpinBox, "dsb_cota_instalacion_celda")
        temperatura_inicial_celda = dialogo.findChild(QDoubleSpinBox, "dsb_temperatura_inicial_celda")
        tk_celda = dialogo.findChild(QDoubleSpinBox, "dsb_tk_celda")
        labelmensaje = dialogo.findChild(QLabel, "label_mensaje")
        boton_guardar = dialogo.findChild(QPushButton, f"btn_guardar_celda")
        cota_superficie_celda.setEnabled(False)
        # cargar data componentes
        componentes = ProyectoController.ctrlObtenerComponentesProyecto(idproyecto)
        if componentes:
            for fila in componentes:
                comboComponente.addItem(str(fila[2]), fila[0])
            comboComponente.setEnabled(True)
        # mostrar data Piezómetro Cuerda
        nombreactual = ""
        idcelda = 0
        datapiezo = CeldaController.ctrlObtenerInfoCelda(idinstrumento)
        if datapiezo:
            idcelda = datapiezo[0]
            comboComponente.setCurrentIndex(comboComponente.findData(idcomponente))
            nombre_celda.setText(str(datapiezo[2]))
            nombreactual = str(datapiezo[2])
            marca_celda.setText(str(datapiezo[3]))
            modelo_celda.setText(str(datapiezo[4]))
            rango_celda.setValue(datapiezo[6])
            coordenada_este_celda.setValue(datapiezo[7])
            coordenada_norte_celda.setValue(datapiezo[8])
            cota_instalacion_celda.setValue(datapiezo[9])
            cota_fundacion_celda.setValue(datapiezo[10])
            frecuencia_inicial_celda.setValue(datapiezo[11])
            cf_celda.setValue(datapiezo[13])
            temperatura_inicial_celda.setValue(datapiezo[12])
            tk_celda.setValue(datapiezo[14])
        def actualizarDatos():
            if nombre_celda and nombre_celda.text().strip():
                componente = comboComponente.currentData()
                nombrezona = comboComponente.currentText()
                nombre = nombre_celda.text()
                celda_data = {
                    "nombre_celda": nombre,
                    "marca_celda": marca_celda.text(),
                    "modelo_celda": modelo_celda.text(),
                    "rango_celda": rango_celda.value(),
                    "frecuencia_inicial": frecuencia_inicial_celda.value(),
                    "cf_celda": cf_celda.value(),
                    "cota_fundacion_celda": cota_fundacion_celda.value(),
                    "coordenada_este_celda": coordenada_este_celda.value(),
                    "coordenada_norte_celda": coordenada_norte_celda.value(),
                    "cota_instalacion_celda": cota_instalacion_celda.value(),
                    "temperatura_inicial_celda": temperatura_inicial_celda.value(),
                    "tk_celda": tk_celda.value(),
                    "idcelda": idcelda,
                    "instrumento": idinstrumento,
                    "componente": componente
                }
                # Llamar al método para guardar en la base de datos
                respuesta = CeldaController.ctrlActualizarCelda(celda_data)
                if respuesta:
                    dialogo.close()
                    if str(idcomponente) == str(componente):
                        TreeCheckbox.actualizarTextoCheckboxEquipo(treewidget, idcomponente, nombregrupo, subgrupo, nombreactual, nombre)
                    else:
                        # Eliminar cota terreno
                        TreeCheckbox.eliminarCheckbox(treewidget, nombregrupo, idinstrumento, subgrupo)
                        # Crear terreno en nuevo componente
                        celda = InterfazController.ctrlListarComponenteCelda(idinstrumento)
                        if celda:
                            if vista == "CELDAS":
                                TreeCheckbox.crearNuevoGrupoCheckboxesDoble(treewidget, nombrezona, componente, idproyecto, nombregrupo, tipogrupo, celda, subgrupo)
                            else:
                                TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidget, nombrezona, componente, idproyecto, nombregrupo, tipogrupo, celda, subgrupo)
                    reiniciarvistas("Celda")
                else:
                    labelmensaje.setText("Error al actualizar la celda.")
            else:
                labelmensaje.setText("El nombre de la celda es obligatorio.")
        # conectar señales
        boton_guardar.clicked.connect(actualizarDatos)
        dialogo.exec()
    
    def eliminar_celda(idproyecto, idinstrumento, nombrecelda, nombregrupo, tipolista, treewidget, reiniciarvistas):
        dlg = QMessageBox()
        dlg.setWindowTitle("Eliminar Celda")
        dlg.setText(f"¿Está seguro eliminar la Celda '{nombrecelda}'?")
        dlg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        dlg.setIcon(QMessageBox.Question)
        result = dlg.exec()
        if result == QMessageBox.Yes:
            respuesta = CeldaController.ctrlEliminarCelda(idinstrumento)
            if respuesta:
                delete = CeldaController.ctrlEliminarCeldaData(idproyecto, respuesta)
                if delete:
                    TreeCheckbox.eliminarCheckbox(treewidget, nombregrupo, idinstrumento, tipolista)
                    reiniciarvistas("Celda")
                else:
                    mostrar_mensaje("Eliminar Celda", "Error al eliminar data de la celda.", "advertencia")
            else:
                mostrar_mensaje("Eliminar Celda", "No se pudo eliminar la celda.", "advertencia")
    
    def cambiar_componente_bloque_celda(idproyecto, idcomponente, parent, treewidget, nombregrupo, tipogrupo, subgrupo, reiniciarvistas):
        dialog = QDialog()
        dialog.setWindowTitle("Mover Celdas Componente")
        layout = QFormLayout(dialog)
        # Campo componente
        label_titulo = QLabel("Componente:")
        combo_componente = QComboBox()
        label_mensaje = QLabel("")
        label_mensaje.setAlignment(Qt.AlignCenter)
        label_mensaje.setStyleSheet("QLabel { color: red; }")
        # cargar data componentes
        componentes = ProyectoController.ctrlObtenerComponentesProyecto(idproyecto)
        if componentes:
            for fila in componentes:
                combo_componente.addItem(str(fila[2]), fila[0])
            combo_componente.setCurrentIndex(combo_componente.findData(idcomponente))
        # Botones
        layout.addRow(label_titulo)
        layout.addRow(combo_componente)
        layout.addRow(label_mensaje)
        button_box = QDialogButtonBox(QDialogButtonBox.Save | QDialogButtonBox.Cancel)
        # Cambiar los textos a español
        button_box.button(QDialogButtonBox.Save).setText("Guardar")
        button_box.button(QDialogButtonBox.Cancel).setText("Cancelar")
        layout.addWidget(button_box)
        # Conectar los botones a las funciones correspondientes
        def actualizarDatos():
            componente = combo_componente.currentData()
            nombrezona = combo_componente.currentText()
            if str(idcomponente) == str(componente):
                dialog.reject()
            else:
                result = False
                hijos_marcados = []
                for i in range(parent.childCount()):
                    hijo = parent.child(i)
                    if hijo is not None and hijo.checkState(0) == Qt.Checked:
                        idinstrumento = hijo.text(2)
                        respuesta = CeldaController.ctrlCambiarCeldaComponente(idinstrumento, componente)
                        if respuesta:
                            hijos_marcados.append(hijo)
                            result = True
                if result:
                    dialog.reject()
                    for hijo in hijos_marcados:
                        idinstrumento = hijo.text(2)
                        TreeCheckbox.eliminarCheckbox(treewidget, nombregrupo, idinstrumento, subgrupo)
                        # Crear prisma en nuevo componente
                        celda = InterfazController.ctrlListarComponenteCelda(idinstrumento)
                        if celda:
                            TreeCheckbox.crearNuevoGrupoCheckboxesSimple(treewidget, nombrezona, componente, idproyecto, nombregrupo, tipogrupo, celda, subgrupo)
                    reiniciarvistas("Celdas de Asentamiento")
                    # Limpiar tabla
                    from views.datos_view import DatosView
                    from modules.datos.vistaDatos import VistaDatos
                    tabla =  DatosView.main.findChild(QTableView, "table_datos")
                    VistaDatos.limpiarTablaDatos(tabla)
                else:
                    label_mensaje.setText("Error al cambiar de componente.")
        button_box.accepted.connect(actualizarDatos)
        button_box.rejected.connect(dialog.reject)
        # Mostrar el diálogo
        dialog.setLayout(layout)
        dialog.exec()
    